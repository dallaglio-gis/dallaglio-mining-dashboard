import streamlit as st
# Set page config as the very first Streamlit command
try:
    st.set_page_config(
        page_title="Combined Mining Dashboard",
        page_icon="⛏️",
        layout="wide",
        initial_sidebar_state="expanded",
    )
except Exception:
    # Ignore if already set by another module/app
    pass

# --- Robust module path setup (works with single or nested folders) ---
import os, sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def add_path(p):
    if p and os.path.isdir(p) and p not in sys.path:
        sys.path.insert(0, p)
        return p

# Try both single-level and nested variants for v1.1.4 and v4
CANDIDATE_PATHS_V114 = [
    os.path.join(BASE_DIR, 'v1.1.4', 'v1.1.4'),
    os.path.join(BASE_DIR, 'v1.1.4'),
]
CANDIDATE_PATHS_V4 = [
    os.path.join(BASE_DIR, 'v4', 'v4'),
    os.path.join(BASE_DIR, 'v4'),
]

FOUND_V114 = [add_path(p) for p in CANDIDATE_PATHS_V114 if add_path(p)]
FOUND_V4   = [add_path(p) for p in CANDIDATE_PATHS_V4   if add_path(p)]

# Optional: create packages so relative imports inside modules work
for pkg_dir in (FOUND_V114 + FOUND_V4):
    if pkg_dir:
        init_file = os.path.join(pkg_dir, '__init__.py')
        try:
            if not os.path.exists(init_file):
                open(init_file, 'a').close()
        except Exception:
            pass

# Store import errors for later display (after st.set_page_config)
IMPORT_ERRORS = []

"""
Combined Mining Dashboard

This Streamlit application consolidates three separate tools into a single
dashboard:

1. **Mining Data Processor** – Extracts, validates and summarizes data from
   mining daily report Excel files.  This logic comes from the original
   *v1.1.4* dashboard and uses the `MiningDataProcessor` class to process
   STOPING, TRAMMING, DEVELOPMENT, HOISTING and BENCHES sheets.

2. **Production Dashboard** – An interactive analytics dashboard for
   exploring production data (stoping, tramming and bench grades).  This
   page exposes the same plots, metrics and tables that were available
   in the standalone `dash3.py` application.

3. **Daily Report Update** – A utility to generate or update geology work
   plan workbooks based on uploaded daily report data.  It leverages the
   logic from `update_daily_report_all_days.py` to create one sheet per
   day of the month and fill it with the appropriate daily, month‑to‑date
   and budget figures.  The output filename can either be specified by
   the user or automatically derived from the month contained in the
   uploaded daily report.

4. **Monthly Stope Performance Updater** – Integrates the standalone
   Monthly Stope Performance app. Upload the MSP workbook and supporting
   files (3‑month rolling forecasts, actual physicals, Mining & Milling
   Plans, Tramming reports, and optional August daily report) to update
   the SUMMARY sheet forecasts/actuals and the PNM/MNP sheet.

To run this app locally, install the required dependencies (streamlit,
openpyxl, pandas, numpy, plotly) and launch with:

```
streamlit run combined_app.py
```

The app uses the August 2025 geology work plan workbook supplied in the
`v4` folder as a template.  When processing a different month, the
template is copied and renamed appropriately.  The update logic then
populates the new workbook with daily sheets for all dates that contain
production data.
"""

import os
import sys
import shutil
import tempfile
from datetime import datetime, date
from io import BytesIO
from typing import Dict, Tuple, Optional, List, Set
import openpyxl



# Now try to import modules and capture any errors
try:
    from mining_processor import MiningDataProcessor
except Exception as e:
    MiningDataProcessor = None
    IMPORT_ERRORS.append(('MiningDataProcessor', e, FOUND_V114))

try:
    import update_daily_report_all_days as udr
    import importlib  # imported here to avoid unused‑import warning when update is absent
except Exception as e:
    udr = None
    IMPORT_ERRORS.append(('update_daily_report_all_days', e, FOUND_V4))

# Try to import config objects
try:
    from config.validation_targets import VALIDATION_TARGETS  # type: ignore
except Exception:
    VALIDATION_TARGETS = {}  # type: ignore
# --- end robust setup ---

import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import re

# --- Data sanitization utilities for Streamlit/Arrow compatibility ---
def _decode_bytes(x):
    if isinstance(x, (bytes, bytearray)):
        try:
            return x.decode('utf-8', errors='replace')
        except Exception:
            return str(x)
    return x

NUMERIC_COL_HINT = re.compile(r'(?i)\b(tonnes|tons|t|kg|gpt|grade|value|qty|quantity|metres|meters|m|au|oz|gold)\b')

def sanitize_for_streamlit(df: pd.DataFrame) -> pd.DataFrame:
    """Clean a DataFrame to be Arrow-serializable in Streamlit.

    - Decode bytes to strings
    - Coerce likely numeric columns to float64, replacing placeholders with 0.0
    - Parse date columns
    - Ensure other object columns are pure strings
    """
    if df is None or df.empty:
        return df

    out = df.copy()

    # First pass: decode all bytes in object columns
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].map(_decode_bytes)

    # Second pass: handle column types
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            out[col] = out[col].fillna(0.0)
            continue

        col_lower = col.lower()
        if NUMERIC_COL_HINT.search(col) or any(h in col_lower for h in ['actual', 'budget', 'mtd']):
            if out[col].dtype == object:
                out[col] = out[col].astype(str)
                out[col] = out[col].replace(['', ' ', '-', '–', '—', 'N/A', 'n/a', 'NA', 'na', 'None', 'none', 'nan', 'NaN'], '0')
                out[col] = out[col].str.replace(r'[^\d\.\-]', '', regex=True)
                out[col] = out[col].replace('', '0')
            out[col] = pd.to_numeric(out[col], errors='coerce').fillna(0.0).astype('float64')
        elif 'date' in col_lower:
            out[col] = pd.to_datetime(out[col], errors='coerce')
        else:
            out[col] = out[col].fillna('').astype(str)
            out[col] = out[col].replace(['nan', 'NaN', 'None', 'none'], '')

    return out

def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode('utf-8')

# Define CURRENT_DIR for compatibility with existing code
CURRENT_DIR = BASE_DIR

# Simple CSS overrides for a cleaner look consistent across pages.
st.markdown(
    """
    <style>
    .metric-container {
        background-color: #f8f9fa;
        border-radius: 5px;
        padding: 10px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.05);
    }
    .metric-label { font-weight: bold; color: #555; }
    .metric-value { font-size: 1.6rem; font-weight: bold; }
    .metric-delta { font-size: 0.9rem; }
    </style>
    """,
    unsafe_allow_html=True,
)

# Display any import errors that occurred during module loading
if IMPORT_ERRORS:
    st.error("⚠️ Module Import Errors Detected")
    for module_name, error, paths_tried in IMPORT_ERRORS:
        st.error(f"Failed to import {module_name}. Tried paths: {paths_tried}")
        st.exception(error)

# ---------------------------------------------------------------------------
# Helper functions reused across pages
# ---------------------------------------------------------------------------
def load_production_data(path: str = 'jan_aug_data_with_bench_grades.xlsx') -> pd.DataFrame:
    """Load the combined January–August production dataset used in the
    Production Dashboard.  The function caches the result to avoid
    re‑reading the file on every rerun.

    Parameters
    ----------
    path : str
        Relative or absolute path to the Excel workbook.

    Returns
    -------
    pd.DataFrame
        A dataframe with properly parsed dates and numeric columns.
    """
    @st.cache_data(show_spinner=False)
    def _loader(p: str) -> pd.DataFrame:
        df = pd.read_excel(p)
        # Convert numeric columns
        num_cols = [
            'Stoping_Actual_t', 'Stoping_Budget_t', 'Tramming_Actual_t', 'Tramming_Budget_t',
            'Stoping_Actual_gpt', 'Stoping_Budget_gpt', 'Tramming_Actual_gpt', 'Tramming_Budget_gpt',
            'Stoping_Actual_kg', 'Stoping_Budget_kg', 'Tramming_Actual_kg', 'Tramming_Budget_kg'
        ]
        for col in num_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        # Parse dates
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'])
        # Final sanitize to ensure Arrow compatibility
        return sanitize_for_streamlit(df)
    return _loader(path)


def derive_output_filename(src_name: str) -> str:
    """Derive a default geology work plan filename from a source
    daily report filename.  The new name has the form
    'Geology Daily Work Plan <Month><Year>.xlsx'.  If a month and
    year cannot be extracted, a generic name using the current date is
    returned.

    Parameters
    ----------
    src_name : str
        The base name of the uploaded daily report file.

    Returns
    -------
    str
        Proposed output filename.
    """
    # Try to use the update script's extractor for consistency
    if udr is not None and hasattr(udr, 'extract_month_year_from_filename'):
        try:
            month_num, year = udr.extract_month_year_from_filename(src_name)
            if 1 <= month_num <= 12:
                month_name = date(year, month_num, 1).strftime('%B')
                return f"Geology Daily Work Plan {month_name}{year}.xlsx"
        except Exception:
            pass
    # Fallback – use today's date if parsing fails
    today = datetime.today()
    month_name = today.strftime('%B')
    return f"Geology Daily Work Plan {month_name}{today.year}.xlsx"


def copy_template_workbook() -> str:
    """Return the path to a fresh copy of the geology work plan template.

    A template workbook is required for the Daily Report Update page.
    The template is the August 2025 work plan included in the v4 folder.
    Each call to this function copies the template into a temporary
    location so that modifications do not affect the original file.

    Returns
    -------
    str
        Path to a writable copy of the template workbook.
    """
    # Try both 'v4' and nested 'v4/v4' to be robust across structures
    candidate_dirs = [
        os.path.join(CURRENT_DIR, 'v4'),
        os.path.join(CURRENT_DIR, 'v4', 'v4'),
    ]
    tried_dirs = []
    template_path = None
    for d in candidate_dirs:
        tried_dirs.append(d)
        if not os.path.isdir(d):
            continue
        for fname in os.listdir(d):
            if fname.startswith('Geology Daily Work Plan') and fname.endswith('.xlsx'):
                template_path = os.path.join(d, fname)
                break
        if template_path:
            break

    if not template_path:
        raise FileNotFoundError(f"Geology Daily Work Plan template not found. Tried: {tried_dirs}")

    # Copy to a temporary file
    tmp_dir = tempfile.gettempdir()
    dest_path = os.path.join(tmp_dir, f"template_copy_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.xlsx")
    shutil.copy(template_path, dest_path)
    return dest_path


# ---------------------------------------------------------------------------
# Monthly Stope Performance helpers (parsing and workbook updates)
# ---------------------------------------------------------------------------
def _normalize_id(id_str: str) -> str:
    """Normalise stope IDs by replacing underscores with spaces, collapsing
    multiple spaces, stripping whitespace and converting to uppercase.

    Many of the source files use slightly different naming conventions
    (e.g. "1L_W28_STOPE" versus "1L W28 STOPE").  This helper reduces those
    inconsistencies so IDs can be matched reliably across different sheets.
    """
    if not id_str:
        return ""
    cleaned = id_str.replace("_", " ").replace("\n", " ")
    cleaned = " ".join(cleaned.split())
    return cleaned.strip().upper()


def _month_add(base_month: int, offset: int) -> int:
    """Return the month index after adding an offset.  Keeps result in 1–12.

    Example: ``_month_add(11, 2) -> 1`` (January of next year).
    """
    return ((base_month - 1 + offset) % 12) + 1


def _extract_month_from_filename(filename: str) -> Optional[int]:
    """Attempt to infer the month index (1–12) from a 3‑month rolling filename.

    The naming convention assumed is "3 Months Rolling_<Month> <Year>.xlsx"
    (case insensitive).  Returns 1 for January through 12 for December,
    or ``None`` if parsing fails.
    """
    m = re.search(r"3\s*Months\s*Rolling[_\s-]*([A-Za-z]+)", filename)
    if not m:
        return None
    month_name = m.group(1).strip().lower()
    try:
        return datetime.strptime(month_name, "%B").month
    except ValueError:
        try:
            return datetime.strptime(month_name, "%b").month
        except ValueError:
            return None


def parse_three_month_rolling(file_bytes: BytesIO, filename: str) -> Optional[Dict[int, Tuple[float, float, float]]]:
    """Parse a 3‑month rolling forecast workbook for forecast data.

    Returns a mapping of month indices (1–12) to tuples ``(tonnes, grade, gold)``.
    The first month is inferred from the filename and subsequent two months
    follow sequentially.  If the workbook structure cannot be parsed,
    ``None`` is returned.
    """
    start_month = _extract_month_from_filename(filename)
    if start_month is None:
        return None
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    sheet = wb.active
    if "MINE PLAN" in wb.sheetnames:
        sheet = wb["MINE PLAN"]
    tonnes: List[float] = []
    grades: List[float] = []
    golds: List[float] = []
    for row in sheet.iter_rows(values_only=True):
        if not row or len(row) < 2 or row[1] is None:
            continue
        label = str(row[1]).strip().lower()
        if label == "tonnes (t)":
            vals: List[float] = []
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    vals.append(float(val))
                if len(vals) >= 3:
                    break
            if vals:
                tonnes = vals
        elif label.startswith("grade"):
            vals = []
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    vals.append(float(val))
                if len(vals) >= 3:
                    break
            if vals:
                grades = vals
        elif label.startswith("3 month rolling") or "gold" in label:
            vals = []
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    vals.append(float(val))
                if len(vals) >= 3:
                    break
            if vals:
                golds = vals
        if tonnes and grades and golds:
            break
    if not golds:
        for row in sheet.iter_rows(values_only=True):
            if row and len(row) > 1 and isinstance(row[1], str) and "gold" in row[1].lower():
                vals: List[float] = []
                for val in row[2:]:
                    if isinstance(val, (int, float)):
                        vals.append(float(val))
                    if len(vals) >= 3:
                        break
                if vals:
                    golds = vals
                    break
    if not tonnes or not grades or not golds:
        return None
    while len(tonnes) < 3:
        tonnes.append(0.0)
    while len(grades) < 3:
        grades.append(0.0)
    while len(golds) < 3:
        golds.append(0.0)
    forecasts: Dict[int, Tuple[float, float, float]] = {}
    for offset in range(3):
        month = _month_add(start_month, offset)
        forecasts[month] = (tonnes[offset], grades[offset], golds[offset])
    return forecasts


def parse_actual_physical(file_bytes: BytesIO) -> Optional[Tuple[float, float, float, int]]:
    """Extract actuals from a single monthly physicals workbook.

    Returns ``(tonnes, grade, gold, month_index)`` or ``None`` if the
    sheet cannot be parsed.  The month index is taken from a datetime near
    the top of the Dashboard sheet.
    """
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    if "Dashboard" not in wb.sheetnames:
        return None
    sheet = wb["Dashboard"]
    month_index: Optional[int] = None
    for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
        for val in row:
            if isinstance(val, datetime):
                month_index = val.month
                break
        if month_index:
            break
    if not month_index:
        return None
    tonne_value = grade_value = gold_value = None
    found_plant = False
    for row in sheet.iter_rows(values_only=True):
        if not found_plant and row and row[0] and "plant physicals" in str(row[0]).lower():
            found_plant = True
            continue
        if found_plant and row and isinstance(row[0], str):
            low = row[0].lower().strip()
            if low.startswith("ore milled"):
                for val in row[1:]:
                    if isinstance(val, (int, float)):
                        tonne_value = float(val)
                        break
            elif low.startswith("mill feed grade"):
                for val in row[1:]:
                    if isinstance(val, (int, float)):
                        grade_value = float(val)
                        break
            elif low.startswith("gold recovered"):
                for val in row[1:]:
                    if isinstance(val, (int, float)):
                        gold_value = float(val)
                        break
        if tonne_value is not None and grade_value is not None and gold_value is not None:
            break
    if tonne_value is None or grade_value is None or gold_value is None:
        return None
    return tonne_value, grade_value, gold_value, month_index


def parse_august_daily_report(file_bytes: BytesIO) -> Optional[Tuple[float, float, float]]:
    """Parse the August daily report and return the Actual MTD values.

    Looks for the ``Tramming`` sheet and reads rows labelled
    ``Trammed (t)``, ``Grade (g/t)``, and ``Gold (kg)`` from the column
    headed “Actual”.
    """
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    if "Tramming" not in wb.sheetnames:
        return None
    sheet = wb["Tramming"]
    tonnes = grade = gold = None
    header_row_index = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and any(isinstance(val, str) and 'mtd' in val.lower() for val in row):
            header_row_index = i
            break
    if header_row_index is None:
        return None
    header_row = [val if isinstance(val, str) else "" for val in sheet[header_row_index]]
    actual_col = None
    for idx, val in enumerate(header_row):
        if isinstance(val, str) and val.lower().strip() == 'actual':
            actual_col = idx + 1
            break
    if actual_col is None:
        actual_col = 4
    for row in sheet.iter_rows(min_row=header_row_index + 1, max_row=header_row_index + 10, values_only=False):
        label_cell = row[1].value if len(row) > 1 else None
        if isinstance(label_cell, str):
            low = label_cell.lower()
            if 'trammed' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    tonnes = float(val)
            elif 'grade' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    grade = float(val)
            elif 'gold' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    gold = float(val)
    if tonnes is None or grade is None or gold is None:
        return None
    return tonnes, grade, gold


def parse_underground_breaking_plan(file_bytes: BytesIO) -> Set[str]:
    """Extract stope IDs from an Underground Breaking Plan workbook."""
    ids: Set[str] = set()
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return ids
    target_name = None
    for name in wb.sheetnames:
        if 'underground breaking plan' in name.lower():
            target_name = name
            break
    if not target_name:
        return ids
    ws = wb[target_name]
    for row in ws.iter_rows(values_only=True):
        if row and len(row) > 1:
            second = row[1]
            if isinstance(second, str) and second.strip().lower() == 'stoping tons':
                stope = row[0]
                if isinstance(stope, str):
                    ids.add(_normalize_id(stope))
    return ids


def parse_tramming_detail(file_bytes: BytesIO) -> Optional[Tuple[int, Dict[str, Tuple[float, float, float, float, float, float]]]]:
    """Parse detailed tramming daily report for stope-level budgets and actuals.

    Returns ``(month_index, data)`` where ``data`` maps normalised stope IDs
    to ``(budget_tonnes, actual_tonnes, budget_grade, actual_grade, budget_gold, actual_gold)``.
    """
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    sheet_name = None
    for name in wb.sheetnames:
        if 'tramming' in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        return None
    ws = wb[sheet_name]
    month_idx: Optional[int] = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        for i, val in enumerate(row):
            if isinstance(val, str) and 'month' in val.lower():
                if i + 1 < len(row) and isinstance(row[i + 1], datetime):
                    month_idx = row[i + 1].month
                    break
        if month_idx:
            break
    if not month_idx:
        return None
    data: Dict[str, Tuple[float, float, float, float, float, float]] = {}
    rows = list(ws.iter_rows(values_only=True))
    n = len(rows)
    i = 0
    while i < n:
        row = rows[i]
        stope_id = None
        if row:
            for cell in row:
                if isinstance(cell, str) and 'stope' in cell.lower():
                    stope_id = _normalize_id(cell)
                    break
        if stope_id:
            budget_tonnes = actual_tonnes = 0.0
            budget_grade = actual_grade = 0.0
            budget_gold = actual_gold = 0.0
            j = i + 1
            while j < n:
                next_row = rows[j]
                if any(isinstance(c, str) and 'stope' in c.lower() for c in next_row if c):
                    break
                if next_row:
                    for k, val in enumerate(next_row):
                        if isinstance(val, str):
                            label = val.strip().lower()
                            if label == 'budget (t)':
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        budget_tonnes = float(x)
                                        break
                            elif label == 'actual (t)':
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        actual_tonnes = float(x)
                                        break
                            elif label in ('budget (g/t)', 'budget (gpt)'):
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        budget_grade = float(x)
                                        break
                            elif label in ('actual (g/t)', 'actual (gpt)'):
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        actual_grade = float(x)
                                        break
                            elif label == 'budget (kg)':
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        budget_gold = float(x)
                                        break
                            elif label == 'actual (kg)':
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        actual_gold = float(x)
                                        break
                j += 1
            data[stope_id] = (budget_tonnes, actual_tonnes, budget_grade, actual_grade, budget_gold, actual_gold)
            i = j
            continue
        i += 1
    return month_idx, data


def update_pnm_mnp_sheet(wb: openpyxl.Workbook,
                         pnm_data: Dict[str, Tuple[float, float, float]],
                         mnp_data: Dict[str, Tuple[float, float, float]],
                         month_idx: int) -> None:
    """Update the ``Stopes PNM & MNP`` sheet with PNM and MNP values for a month."""
    if 'Stopes PNM & MNP' not in wb.sheetnames:
        return
    ws = wb['Stopes PNM & MNP']

    def build_month_map(start_row: int) -> Dict[int, int]:
        month_map: Dict[int, int] = {}
        header_row = ws[start_row + 1]
        for col_idx, cell in enumerate(header_row, start=1):
            if isinstance(cell.value, datetime):
                month_map[cell.value.month] = col_idx
        return month_map

    pnm_start = mnp_start = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and len(row) > 1 and isinstance(row[1], str):
            val = row[1].strip().lower()
            if val == 'planned not mined' and pnm_start is None:
                pnm_start = idx
            elif val == 'mined not planned' and mnp_start is None:
                mnp_start = idx
        if pnm_start and mnp_start:
            break
    if pnm_start is None or mnp_start is None:
        return

    pnm_month_map = build_month_map(pnm_start + 1)
    mnp_month_map = build_month_map(mnp_start + 1)
    pnm_col_start = pnm_month_map.get(month_idx)
    mnp_col_start = mnp_month_map.get(month_idx)
    if not pnm_col_start and not mnp_col_start:
        return

    def find_stope_rows(section_start: int) -> Tuple[int, int]:
        start = section_start + 3
        end = start - 1
        for r in range(start, ws.max_row + 1):
            val = ws.cell(row=r, column=2).value
            if val is None:
                break
            if isinstance(val, str) and val.strip().lower() == 'total':
                end = r - 1
                break
        return start, end

    pnm_stope_start, pnm_stope_end = find_stope_rows(pnm_start)
    mnp_stope_start, mnp_stope_end = find_stope_rows(mnp_start)

    def get_or_create_row(section_start: int, stope_start: int, stope_end: int, stope_id: str) -> int:
        for r in range(stope_start, stope_end + 1):
            val = ws.cell(row=r, column=2).value
            if isinstance(val, str) and _normalize_id(val) == stope_id:
                return r
        insert_row = stope_end + 1
        ws.insert_rows(insert_row)
        ws.cell(row=insert_row, column=2, value=stope_id)
        return insert_row

    if pnm_col_start:
        for stope_id, (diff_tonnes, diff_grade, diff_gold) in pnm_data.items():
            row_idx = get_or_create_row(pnm_start, pnm_stope_start, pnm_stope_end, stope_id)
            ws.cell(row=row_idx, column=pnm_col_start, value=diff_tonnes)
            ws.cell(row=row_idx, column=pnm_col_start + 1, value=diff_grade)
            ws.cell(row=row_idx, column=pnm_col_start + 2, value=diff_gold)
            if row_idx > pnm_stope_end:
                pnm_stope_end = row_idx
        total_row = pnm_stope_end + 1
        tot_tonnes = 0.0
        weighted_grade_sum = 0.0
        tot_gold = 0.0
        for r in range(pnm_stope_start, pnm_stope_end + 1):
            v_t = ws.cell(row=r, column=pnm_col_start).value or 0
            v_g = ws.cell(row=r, column=pnm_col_start + 1).value or 0
            v_au = ws.cell(row=r, column=pnm_col_start + 2).value or 0
            try:
                v_t = float(v_t)
            except Exception:
                v_t = 0.0
            try:
                v_g = float(v_g)
            except Exception:
                v_g = 0.0
            try:
                v_au = float(v_au)
            except Exception:
                v_au = 0.0
            tot_tonnes += v_t
            weighted_grade_sum += v_t * v_g
            tot_gold += v_au
        ws.cell(row=total_row, column=pnm_col_start, value=tot_tonnes)
        avg_grade = weighted_grade_sum / tot_tonnes if tot_tonnes else 0
        ws.cell(row=total_row, column=pnm_col_start + 1, value=avg_grade)
        ws.cell(row=total_row, column=pnm_col_start + 2, value=tot_gold)

    if mnp_col_start:
        for stope_id, (act_tonnes, act_grade, act_gold) in mnp_data.items():
            row_idx = get_or_create_row(mnp_start, mnp_stope_start, mnp_stope_end, stope_id)
            ws.cell(row=row_idx, column=mnp_col_start, value=act_tonnes)
            ws.cell(row=row_idx, column=mnp_col_start + 1, value=act_grade)
            ws.cell(row=row_idx, column=mnp_col_start + 2, value=act_gold)
            if row_idx > mnp_stope_end:
                mnp_stope_end = row_idx
        total_row = mnp_stope_end + 1
        tot_tonnes = 0.0
        weighted_grade_sum = 0.0
        tot_gold = 0.0
        for r in range(mnp_stope_start, mnp_stope_end + 1):
            v_t = ws.cell(row=r, column=mnp_col_start).value or 0
            v_g = ws.cell(row=r, column=mnp_col_start + 1).value or 0
            v_au = ws.cell(row=r, column=mnp_col_start + 2).value or 0
            try:
                v_t = float(v_t)
            except Exception:
                v_t = 0.0
            try:
                v_g = float(v_g)
            except Exception:
                v_g = 0.0
            try:
                v_au = float(v_au)
            except Exception:
                v_au = 0.0
            tot_tonnes += v_t
            weighted_grade_sum += v_t * v_g
            tot_gold += v_au
        ws.cell(row=total_row, column=mnp_col_start, value=tot_tonnes)
        avg_grade = weighted_grade_sum / tot_tonnes if tot_tonnes else 0
        ws.cell(row=total_row, column=mnp_col_start + 1, value=avg_grade)
        ws.cell(row=total_row, column=mnp_col_start + 2, value=tot_gold)


def update_msp_workbook(msp_bytes: BytesIO,
                        forecasts: Dict[int, Tuple[float, float, float]],
                        actuals: Dict[int, Tuple[float, float, float]]) -> BytesIO:
    """Update the MSP SUMMARY sheet with forecasts and actuals and return bytes."""
    wb = openpyxl.load_workbook(msp_bytes)
    if "SUMMARY" not in wb.sheetnames:
        raise ValueError("SUMMARY sheet not found in MSP workbook")
    ws = wb["SUMMARY"]
    forecast_start = actual_start = None
    for row_idx in range(1, ws.max_row + 1):
        row_label = ws.cell(row=row_idx, column=1).value
        row_type = ws.cell(row=row_idx, column=2).value
        if forecast_start is None and row_label == 'SHORT-TERM FORECAST' and row_type == 'Tonnes (t)':
            forecast_start = row_idx
        if actual_start is None and row_label == 'ACTUAL' and row_type == 'Tonnes (t)':
            actual_start = row_idx
        if forecast_start and actual_start:
            break
    if forecast_start is None or actual_start is None:
        raise ValueError("Could not locate forecast or actual rows in SUMMARY")
    forecast_rows = (forecast_start, forecast_start + 1, forecast_start + 2)
    actual_rows = (actual_start, actual_start + 1, actual_start + 2)
    jan_col = 3
    for month_idx, (t, g, au) in forecasts.items():
        col = jan_col + month_idx - 1
        ws.cell(row=forecast_rows[0], column=col).value = t
        ws.cell(row=forecast_rows[1], column=col).value = g
        ws.cell(row=forecast_rows[2], column=col).value = au
    for month_idx, (t, g, au) in actuals.items():
        col = jan_col + month_idx - 1
        ws.cell(row=actual_rows[0], column=col).value = t
        ws.cell(row=actual_rows[1], column=col).value = g
        ws.cell(row=actual_rows[2], column=col).value = au
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------------------------------------------------------------------------
# Mining Data Processor page
# ---------------------------------------------------------------------------
def run_mining_processor_page():
    """Display the Mining Data Processor interface.  This page allows users
    to upload a mining daily report, select which sheets to process,
    optionally validate against target metrics and produce summary
    reports.  The logic here mirrors the original v1.1.4 Streamlit
    application but is confined to a single function to coexist with
    other pages.
    """
    st.markdown("## ⛏️ Mining Data Processor")
    st.markdown("Process your mining daily report and extract structured data from the STOPING, TRAMMING, DEVELOPMENT, HOISTING and BENCHES sheets.")

    # Early exit if the processor module cannot be imported
    if MiningDataProcessor is None:
        st.warning("The Mining Data Processor module could not be loaded. Please ensure that the v1.1.4 folder is present.")
        return

    # Sidebar options for this page
    with st.sidebar:
        st.header("Processing Configuration")
        uploaded_file = st.file_uploader(
            "Upload Excel Daily Report",
            type=['xlsx', 'xls'],
            help="Select your mining daily report Excel file for extraction."
        )
        st.markdown("Select which sheets to process:")
        sheet_options = ['STOPING', 'TRAMMING', 'DEVELOPMENT', 'HOISTING', 'BENCHES']
        selected_sheets = []
        for sheet in sheet_options:
            if st.checkbox(f"{sheet}", value=True, key=f"mp_sheet_{sheet}"):
                selected_sheets.append(sheet)
        st.markdown("Processing options:")
        include_validation = st.checkbox("Enable Validation Against Targets", value=True, key="mp_validate")
        include_visualization = st.checkbox("Generate Data Visualizations", value=True, key="mp_vis")
        create_summary_report = st.checkbox("Create Summary Report", value=True, key="mp_summary")
        output_dir = st.text_input("Output Directory", value="outputs", key="mp_outdir")
        # Debug toggle
        st.checkbox("Show Debug Column Types", value=False, key="mp_show_dtypes")

    if uploaded_file is None:
        # Show a welcome message when no file is uploaded
        st.info("Upload an Excel file to begin processing.")
        return

    # Save uploaded file to a temporary location.  The original filename is
    # preserved so that downstream functions (like month extraction) can
    # infer information from it.  Using NamedTemporaryFile ensures proper
    # cleanup on program exit.
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[-1], prefix=os.path.splitext(uploaded_file.name)[0] + '_') as tmp_file:
        tmp_file.write(uploaded_file.getvalue())
        temp_file_path = tmp_file.name

    # Create a processor instance
    processor = MiningDataProcessor(output_dir)

    # Validate the file before processing
    st.subheader("File Information")
    file_col1, file_col2 = st.columns(2)
    with file_col1:
        st.metric("Filename", uploaded_file.name)
    with file_col2:
        st.metric("File Size", f"{uploaded_file.size / 1024:.1f} KB")

    st.subheader("File Validation")
    with st.spinner("Validating Excel file..."):
        is_valid, missing_sheets = processor.validate_excel_file(temp_file_path)
    if is_valid:
        st.success("All required sheets found.")
    else:
        st.error(f"Missing sheets: {missing_sheets}")
        return

    # Allow user to start processing
    st.subheader("Data Processing")
    if st.button("Start Processing", type="primary"):
        if not selected_sheets:
            st.error("Please select at least one sheet to process.")
        else:
            results = process_all_sheets_wrapper(processor, temp_file_path, selected_sheets)
            display_mining_results(results, include_visualization)


def process_all_sheets_wrapper(processor: MiningDataProcessor, file_path: str, selected_sheets: list) -> dict:
    """Wrapper around the MiningDataProcessor.process_all_sheets call.  It
    collects the results and attaches a processing timestamp for later
    display.  Any exceptions are caught and returned in the result
    dictionary.
    """
    try:
        results = processor.process_all_sheets(file_path, selected_sheets)
        return results
    except Exception as e:
        return {'error': str(e)}


def display_mining_results(results: dict, include_visualization: bool):
    """Render the results of the MiningDataProcessor on the page.  If an
    error occurred, it is displayed to the user.  Otherwise the
    per‑sheet tables, validation summaries and optional charts are
    shown.
    """
    if 'error' in results:
        st.error(f"Processing error: {results['error']}")
        return

    st.header("Processing Results")

    overall = results.get('overall_summary', {})
    summary_cols = st.columns(4)
    with summary_cols[0]:
        st.metric("Sheets Processed", f"{overall.get('successful_sheets', 0)}/{overall.get('total_sheets_requested', 0)}")
    with summary_cols[1]:
        st.metric("Total Records", overall.get('total_records', 0))
    with summary_cols[2]:
        st.metric("Processing Time", results.get('processing_time', 'N/A'))
    with summary_cols[3]:
        st.metric("Output Directory", overall.get('output_directory', 'N/A'))

    # Iterate over each sheet's results
    for sheet_type, sheet_result in results.get('sheets_processed', {}).items():
        with st.expander(f"{sheet_type} Results", expanded=True):
            if sheet_result.get('success'):
                st.success(f"{sheet_type} processed successfully")
                info_cols = st.columns(2)
                with info_cols[0]:
                    st.metric("Records Extracted", len(sheet_result['data']))
                with info_cols[1]:
                    if 'output_file' in sheet_result:
                        st.text(f"Output: {os.path.basename(sheet_result['output_file'])}")
                # Validation results
                if sheet_result.get('validation'):
                    display_validation_results(sheet_result['validation'])
                # Preview data
                st.subheader(f"{sheet_type} Data Preview")
                df_raw = sheet_result['data']
                df = sanitize_for_streamlit(df_raw)
                if not df.empty:
                    # Optional debug: dtypes and sample unique values
                    if st.session_state.get('mp_show_dtypes', False):
                        with st.expander("Debug: Column Types and Samples", expanded=False):
                            st.write("Dtypes:")
                            st.write(pd.DataFrame(df.dtypes.astype(str), columns=['dtype']))
                            # Show unique samples for object columns
                            obj_cols = [c for c in df.columns if df[c].dtype == object]
                            for c in obj_cols[:8]:  # limit to 8 columns for brevity
                                uniq = df[c].dropna().unique()[:10]
                                st.write({c: uniq})
                    st.dataframe(df.head(10), use_container_width=True)
                    csv_bytes = to_csv_bytes(df)
                    st.download_button(
                        label=f"Download {sheet_type} Data",
                        data=csv_bytes,
                        file_name=f"{sheet_type.lower()}_data.csv",
                        mime="text/csv",
                        key=f"mp_dl_{sheet_type.lower()}"
                    )
                    st.caption(f"File size: {len(csv_bytes):,} bytes")
                    if include_visualization:
                        display_visualizations(df, sheet_type)
                # Additional average grades for benches
                if sheet_type == 'BENCHES' and 'data_avg' in sheet_result:
                    st.subheader("Average Grades Data")
                    avg_df_raw = sheet_result['data_avg']
                    avg_df = sanitize_for_streamlit(avg_df_raw)
                    st.dataframe(avg_df.head(10), use_container_width=True)
                    avg_bytes = to_csv_bytes(avg_df)
                    st.download_button(
                        label="Download Average Grades",
                        data=avg_bytes,
                        file_name="benches_average_grades.csv",
                        mime="text/csv",
                        key="mp_dl_benches_avg",
                    )
                    st.caption(f"File size: {len(avg_bytes):,} bytes")
            else:
                st.error(f"{sheet_type} processing failed: {sheet_result.get('error', 'Unknown error')}")

    # Bulk download of all results
    st.subheader("Bulk Download")
    # Always show the ZIP download button if results exist
    create_bulk_download(results)


def create_bulk_download(results: dict):
    """Package all extracted sheet data into a single ZIP archive for download.
    Each CSV is written into the archive in memory and the result is
    offered to the user via a download button.
    """
    import zipfile
    import io
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for sheet_type, sheet_result in results.get('sheets_processed', {}).items():
            if sheet_result.get('success') and not sheet_result['data'].empty:
                clean = sanitize_for_streamlit(sheet_result['data'])
                csv_bytes = to_csv_bytes(clean)
                zf.writestr(f"{sheet_type.lower()}_data.csv", csv_bytes)
                if sheet_type == 'BENCHES' and 'data_avg' in sheet_result:
                    avg_clean = sanitize_for_streamlit(sheet_result['data_avg'])
                    avg_bytes = to_csv_bytes(avg_clean)
                    zf.writestr("benches_average_grades.csv", avg_bytes)

    zip_buffer.seek(0)
    zip_bytes = zip_buffer.getvalue()
    st.download_button(
        label="Download ZIP Archive",
        data=zip_bytes,
        file_name=f"mining_extraction_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
        mime="application/zip",
        key="mp_zip_download",
    )
    st.caption(f"ZIP size: {len(zip_bytes):,} bytes")


# ---------------------------------------------------------------------------
# Validation and Visualization helpers (Mining Processor page)
# ---------------------------------------------------------------------------
def display_validation_results(validation: dict):
    """Render validation summary and per-metric results.

    Expected format:
    {
      'passed': int,
      'failed': int,
      'details': {
         metric_name: {'passed': bool, 'target': num, 'actual': num, 'diff_percentage': float}
      }
    }
    """
    try:
        st.subheader("Validation Summary")
        c1, c2 = st.columns(2)
        with c1:
            st.metric("Tests Passed", int(validation.get('passed', 0)))
        with c2:
            st.metric("Tests Failed", int(validation.get('failed', 0)))

        details = validation.get('details', {})
        if isinstance(details, dict) and details:
            for metric, result in details.items():
                status = "✅" if result.get('passed') else "❌"
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.text(f"{status} {metric}")
                with col2:
                    tgt = result.get('target', '')
                    if isinstance(tgt, (int, float, np.floating)):
                        st.text(f"Target: {tgt:.2f}")
                    else:
                        st.text(f"Target: {tgt}")
                with col3:
                    act = result.get('actual', '')
                    if isinstance(act, (int, float, np.floating)):
                        st.text(f"Actual: {act:.2f}")
                    else:
                        st.text(f"Actual: {act}")
                with col4:
                    dp = result.get('diff_percentage', None)
                    if isinstance(dp, (int, float, np.floating)):
                        st.text(f"Diff: {dp*100:.1f}%")
                    else:
                        st.text("Diff: —")
        else:
            st.info("No detailed validation results available.")
    except Exception as e:
        st.warning(f"Could not display validation results: {e}")


def display_visualizations(df: pd.DataFrame, sheet_type: str):
    """Lightweight visuals for extracted data. Uses sanitized df."""
    st.subheader(f"📊 {sheet_type} Visualizations")
    df_clean = df.copy()

    # Helper: find two numeric columns
    def pick_two_numeric(data: pd.DataFrame):
        num_cols = [c for c in data.columns if pd.api.types.is_numeric_dtype(data[c])]
        return num_cols[:2]

    try:
        # Time series if Date present
        if 'Date' in df_clean.columns:
            nc = pick_two_numeric(df_clean)
            if len(nc) >= 1:
                daily = df_clean.groupby('Date')[nc].sum().reset_index()
                fig = px.line(daily, x='Date', y=nc, title=f"Daily {sheet_type} Metrics")
                fig.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
                st.plotly_chart(fig, use_container_width=True)

        # Category breakdown by ID/Stope if present
        id_col = 'ID' if 'ID' in df_clean.columns else ('Stope_ID' if 'Stope_ID' in df_clean.columns else None)
        if id_col is not None:
            nc = pick_two_numeric(df_clean)
            if nc:
                grp = df_clean.groupby(id_col)[nc[0]].sum().sort_values(ascending=False).head(10)
                fig_bar = px.bar(x=grp.index, y=grp.values, title=f"Top 10 by {id_col}")
                fig_bar.update_layout(xaxis_title=id_col, yaxis_title=nc[0])
                st.plotly_chart(fig_bar, use_container_width=True)

        # Special cases
        if sheet_type.upper() == 'HOISTING' and {'Source', 'Value'}.issubset(df_clean.columns):
            src_sum = df_clean.groupby('Source')['Value'].sum().sort_values(ascending=False)
            fig_pie = px.pie(values=src_sum.values, names=src_sum.index, title="Hoisting by Source")
            st.plotly_chart(fig_pie, use_container_width=True)

        if sheet_type.upper() == 'BENCHES' and 'AU' in df_clean.columns:
            au = df_clean['AU']
            au = au[au > 0] if pd.api.types.is_numeric_dtype(au) else pd.to_numeric(au, errors='coerce').dropna()
            if len(au) > 0:
                fig_hist = px.histogram(x=au, nbins=30, title="Gold Grade Distribution (AU)")
                fig_hist.update_layout(xaxis_title="AU (g/t)", yaxis_title="Frequency")
                st.plotly_chart(fig_hist, use_container_width=True)
    except Exception as e:
        st.info(f"Visualizations not available: {e}")


# ---------------------------------------------------------------------------
# Production Dashboard page
# ---------------------------------------------------------------------------
def run_production_dashboard_page():
    """Render the production analytics dashboard.  This page closely
    follows the layout of the original `dash3.py` dashboard.  Users can
    filter by date range and stope ID and explore key metrics, trends
    and comparisons across stoping, tramming and bench data.
    """
    st.markdown("## 📊 Production Dashboard")
    st.markdown("Interactive analysis of production data across stoping, tramming and benches.")

    # Load the data once.  The dataset must reside in the same directory as this script.
    data_path = os.path.join(CURRENT_DIR, 'jan_aug_data_with_bench_grades.xlsx')
    if not os.path.exists(data_path):
        st.error("Dataset jan_aug_data_with_bench_grades.xlsx not found.")
        return
    df = load_production_data(data_path)

    # Sidebar filters for the dashboard page
    with st.sidebar:
        st.header("Filters")
        min_date = df['Date'].min().date()
        max_date = df['Date'].max().date()
        start_date = st.date_input("Start Date", min_value=min_date, max_value=max_date, value=min_date, key="pd_start")
        end_date = st.date_input("End Date", min_value=min_date, max_value=max_date, value=max_date, key="pd_end")
        # Stope ID selector
        all_stopes = sorted(df['Stope_ID'].dropna().unique())
        default_stopes = all_stopes[:5] if len(all_stopes) >= 5 else all_stopes
        selected_stopes = st.multiselect("Select Stopes", options=all_stopes, default=default_stopes, key="pd_stopes")
        # Export button
        st.markdown("---")
        st.markdown("### Export Data")
        filtered_df_tmp = df[(df['Date'].dt.date >= start_date) & (df['Date'].dt.date <= end_date) & (df['Stope_ID'].isin(selected_stopes))]
        csv = to_csv_bytes(filtered_df_tmp)
        st.download_button(
            label="Download Filtered Data",
            data=csv,
            file_name="filtered_production_data.csv",
            mime="text/csv",
            key="pd_download_filtered",
        )
        st.caption(f"File size: {len(csv):,} bytes")

    # Apply filters to data
    filtered_df = df[(df['Date'].dt.date >= start_date) & (df['Date'].dt.date <= end_date) & (df['Stope_ID'].isin(selected_stopes))]

    # Main header
    st.markdown("### Daily Production Metrics")
    metric_cols = st.columns(3)
    with metric_cols[0]:
        total_actual_tonnes = filtered_df['Stoping_Actual_t'].sum()
        total_budget_tonnes = filtered_df['Stoping_Budget_t'].sum()
        tonnes_variance = ((total_actual_tonnes / total_budget_tonnes) - 1) * 100 if total_budget_tonnes else 0
        st.metric("Total Stoping Tonnes", f"{total_actual_tonnes:,.0f} t", delta=f"{tonnes_variance:.1f}% vs Budget")
    with metric_cols[1]:
        avg_actual_grade = filtered_df['Stoping_Actual_gpt'].mean()
        avg_budget_grade = filtered_df['Stoping_Budget_gpt'].mean()
        grade_variance = ((avg_actual_grade / avg_budget_grade) - 1) * 100 if avg_budget_grade else 0
        st.metric("Average Stoping Grade", f"{avg_actual_grade:.2f} g/t", delta=f"{grade_variance:.1f}% vs Budget")
    with metric_cols[2]:
        total_actual_gold = filtered_df['Stoping_Actual_kg'].sum()
        total_budget_gold = filtered_df['Stoping_Budget_kg'].sum()
        gold_variance = ((total_actual_gold / total_budget_gold) - 1) * 100 if total_budget_gold else 0
        st.metric("Total Gold Production", f"{total_actual_gold:.2f} kg", delta=f"{gold_variance:.1f}% vs Budget")

    # Tabs for different analyses
    tabs = st.tabs([
        "Production Trends", "Stope Performance", "Grade Analysis",
        "Actual vs Budget", "Stope Drill‑Down", "Stoping vs Tramming"
    ])

    # Tab 1: Production Trends
    with tabs[0]:
        st.markdown("#### Production Trends")
        # Prepare daily aggregated data
        daily_data = filtered_df.groupby('Date').agg({
            'Stoping_Actual_t': 'sum', 'Stoping_Budget_t': 'sum',
            'Stoping_Actual_gpt': 'mean', 'Stoping_Budget_gpt': 'mean',
            'Stoping_Actual_kg': 'sum', 'Stoping_Budget_kg': 'sum'
        }).reset_index()
        cols = st.columns(2)
        # Tonnes trend
        tonnes_choices = ['Stoping_Actual_t', 'Stoping_Budget_t']
        sel_tonnes = st.multiselect("Select Tonnes Metrics", options=tonnes_choices, default=tonnes_choices, key="pd_tonnes_metrics")
        if sel_tonnes:
            fig_tonnes = px.line(daily_data, x='Date', y=sel_tonnes, labels={'value': 'Tonnes', 'variable': 'Metric'}, title='Daily Stoping Tonnes')
            fig_tonnes.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_tonnes, use_container_width=True)
        else:
            st.warning("Select at least one metric for Tonnes.")
        # Gold trend
        gold_choices = ['Stoping_Actual_kg', 'Stoping_Budget_kg']
        sel_gold = st.multiselect("Select Gold Metrics", options=gold_choices, default=gold_choices, key="pd_gold_metrics")
        if sel_gold:
            fig_gold = px.line(daily_data, x='Date', y=sel_gold, labels={'value': 'Gold (kg)', 'variable': 'Metric'}, title='Daily Gold Production')
            fig_gold.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_gold, use_container_width=True)
        else:
            st.warning("Select at least one metric for Gold.")
        # Grade trend
        grade_choices = ['Stoping_Actual_gpt', 'Stoping_Budget_gpt']
        sel_grade = st.multiselect("Select Grade Metrics", options=grade_choices, default=grade_choices, key="pd_grade_metrics")
        if sel_grade:
            fig_grade = px.line(daily_data, x='Date', y=sel_grade, labels={'value': 'Grade (g/t)', 'variable': 'Metric'}, title='Daily Stoping Grade')
            fig_grade.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_grade, use_container_width=True)
        else:
            st.warning("Select at least one metric for Grade.")
        # 3D scatter plot
        fig_3d = px.scatter_3d(
            daily_data,
            x='Stoping_Actual_t', y='Stoping_Actual_gpt', z='Stoping_Actual_kg',
            color='Stoping_Actual_kg', size='Stoping_Actual_t',
            color_continuous_scale=px.colors.sequential.Viridis,
            title='3D Relationship: Tonnes, Grade & Gold Production'
        )
        fig_3d.update_layout(height=450)
        st.plotly_chart(fig_3d, use_container_width=True)

    # Tab 2: Stope Performance
    with tabs[1]:
        st.markdown("#### Stope Performance")
        perf = filtered_df.groupby('Stope_ID').agg({
            'Stoping_Actual_t': 'sum', 'Stoping_Budget_t': 'sum',
            'Stoping_Actual_gpt': 'mean', 'Stoping_Budget_gpt': 'mean',
            'Stoping_Actual_kg': 'sum', 'Stoping_Budget_kg': 'sum'
        }).reset_index()
        perf['Tonnes_Variance'] = ((perf['Stoping_Actual_t'] / perf['Stoping_Budget_t']) - 1) * 100
        perf['Grade_Variance'] = ((perf['Stoping_Actual_gpt'] / perf['Stoping_Budget_gpt']) - 1) * 100
        perf['Gold_Variance'] = ((perf['Stoping_Actual_kg'] / perf['Stoping_Budget_kg']) - 1) * 100
        # Heatmap
        heatmap_options = ['Tonnes_Variance', 'Grade_Variance', 'Gold_Variance']
        sel_heatmap = st.multiselect("Metrics for Heatmap", options=heatmap_options, default=heatmap_options, key="pd_heatmap")
        if sel_heatmap:
            heatmap_data = perf[['Stope_ID'] + sel_heatmap].set_index('Stope_ID')
            fig_heat = px.imshow(
                heatmap_data.T,
                text_auto='.1f',
                aspect='auto',
                color_continuous_scale=[[0, 'red'], [0.5, 'white'], [1, 'green']],
                color_continuous_midpoint=0,
                title='Performance Variance by Stope (%)',
                labels=dict(x='Stope ID', y='Metric', color='Variance %')
            )
            fig_heat.update_layout(height=450)
            st.plotly_chart(fig_heat, use_container_width=True)
        else:
            st.warning("Select at least one metric for the heatmap.")
        # Detailed table
        st.markdown("#### Detailed Performance")
        disp_cols = ['Stope_ID', 'Stoping_Actual_t', 'Stoping_Budget_t', 'Tonnes_Variance',
                     'Stoping_Actual_gpt', 'Stoping_Budget_gpt', 'Grade_Variance',
                     'Stoping_Actual_kg', 'Stoping_Budget_kg', 'Gold_Variance']
        table = perf[disp_cols].rename(columns={
            'Stope_ID': 'Stope ID',
            'Stoping_Actual_t': 'Actual Tonnes', 'Stoping_Budget_t': 'Budget Tonnes', 'Tonnes_Variance': 'Tonnes Var %',
            'Stoping_Actual_gpt': 'Actual Grade (g/t)', 'Stoping_Budget_gpt': 'Budget Grade (g/t)', 'Grade_Variance': 'Grade Var %',
            'Stoping_Actual_kg': 'Actual Gold (kg)', 'Stoping_Budget_kg': 'Budget Gold (kg)', 'Gold_Variance': 'Gold Var %'
        })
        st.dataframe(table.style.format({
            'Actual Tonnes': '{:,.0f}', 'Budget Tonnes': '{:,.0f}', 'Tonnes Var %': '{:+.1f}%',
            'Actual Grade (g/t)': '{:.2f}', 'Budget Grade (g/t)': '{:.2f}', 'Grade Var %': '{:+.1f}%',
            'Actual Gold (kg)': '{:.2f}', 'Budget Gold (kg)': '{:.2f}', 'Gold Var %': '{:+.1f}%'
        }), height=300)

    # Tab 3: Grade Analysis
    with tabs[2]:
        st.markdown("#### Grade Analysis")
        # Bench vs stoping grade comparison
        cols = st.columns(2)
        with cols[0]:
            avg_bench_grade = filtered_df['BENCHES_Avg_Grade'].mean()
            avg_stoping_grade = filtered_df['Stoping_Actual_gpt'].mean()
            diff_pct = ((avg_stoping_grade - avg_bench_grade) / avg_bench_grade) * 100 if avg_bench_grade else 0
            st.metric("Average Bench Grade", f"{avg_bench_grade:.2f} g/t")
            st.metric("Average Stoping Grade", f"{avg_stoping_grade:.2f} g/t", delta=f"{diff_pct:.1f}% vs Bench")
        with cols[1]:
            scatter_df = filtered_df[['Date', 'Stope_ID', 'BENCHES_Avg_Grade', 'Stoping_Actual_gpt']].dropna()
            fig_scatter = px.scatter(
                scatter_df,
                x='BENCHES_Avg_Grade', y='Stoping_Actual_gpt', color='Stope_ID',
                hover_data=['Date'],
                title='Bench Grade vs Stoping Grade',
                labels={'BENCHES_Avg_Grade': 'Bench Grade (g/t)', 'Stoping_Actual_gpt': 'Stoping Grade (g/t)'}
            )
            min_val = min(scatter_df['BENCHES_Avg_Grade'].min(), scatter_df['Stoping_Actual_gpt'].min())
            max_val = max(scatter_df['BENCHES_Avg_Grade'].max(), scatter_df['Stoping_Actual_gpt'].max())
            fig_scatter.add_trace(
                go.Scatter(
                    x=[min_val, max_val], y=[min_val, max_val], mode='lines', line=dict(color='black', dash='dash'), name='Perfect Correlation'
                )
            )
            st.plotly_chart(fig_scatter, use_container_width=True)
        # Grade distribution by stope
        st.markdown("##### Grade Distribution by Stope")
        fig_box = px.box(filtered_df.dropna(subset=['Stoping_Actual_gpt']), x='Stope_ID', y='Stoping_Actual_gpt', color='Stope_ID', title='Stoping Grade Distribution by Stope')
        fig_box.update_layout(showlegend=False)
        st.plotly_chart(fig_box, use_container_width=True)
        # Grade trends over time
        st.markdown("##### Grade Trends Over Time")
        grade_trends = filtered_df.groupby('Date').agg({
            'BENCHES_Avg_Grade': 'mean', 'Stoping_Actual_gpt': 'mean', 'Stoping_Budget_gpt': 'mean'
        }).reset_index()
        trend_options = ['BENCHES_Avg_Grade', 'Stoping_Actual_gpt', 'Stoping_Budget_gpt']
        sel_trends = st.multiselect("Select Metrics", options=trend_options, default=trend_options, key="pd_trend_metrics")
        if sel_trends:
            fig_trend = px.line(grade_trends, x='Date', y=sel_trends, labels={'value': 'Grade (g/t)', 'variable': 'Metric'}, title='Grade Trends Over Time')
            fig_trend.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_trend, use_container_width=True)
        else:
            st.warning("Select at least one metric for the grade trends.")

    # Tab 4: Actual vs Budget
    with tabs[3]:
        st.markdown("#### Actual vs Budget Comparison")
        perf = perf  # reuse from earlier
        cols = st.columns(2)
        with cols[0]:
            fig_scatter = px.scatter(perf, x='Stoping_Budget_t', y='Stoping_Actual_t', color='Stope_ID', title='Actual vs Budget Tonnes', labels={'Stoping_Budget_t': 'Budget Tonnes', 'Stoping_Actual_t': 'Actual Tonnes'})
            min_val = min(perf['Stoping_Budget_t'].min(), perf['Stoping_Actual_t'].min())
            max_val = max(perf['Stoping_Budget_t'].max(), perf['Stoping_Actual_t'].max())
            fig_scatter.add_trace(go.Scatter(x=[min_val, max_val], y=[min_val, max_val], mode='lines', line=dict(color='black', dash='dash'), name='Perfect Correlation'))
            st.plotly_chart(fig_scatter, use_container_width=True)
        with cols[1]:
            fig_scatter2 = px.scatter(perf, x='Stoping_Budget_kg', y='Stoping_Actual_kg', color='Stope_ID', title='Actual vs Budget Gold', labels={'Stoping_Budget_kg': 'Budget Gold (kg)', 'Stoping_Actual_kg': 'Actual Gold (kg)'})
            min_val = min(perf['Stoping_Budget_kg'].min(), perf['Stoping_Actual_kg'].min())
            max_val = max(perf['Stoping_Budget_kg'].max(), perf['Stoping_Actual_kg'].max())
            fig_scatter2.add_trace(go.Scatter(x=[min_val, max_val], y=[min_val, max_val], mode='lines', line=dict(color='black', dash='dash'), name='Perfect Correlation'))
            st.plotly_chart(fig_scatter2, use_container_width=True)
        # Cumulative charts
        st.markdown("##### Cumulative Production")
        daily_sorted = daily_data.sort_values('Date')
        daily_sorted['Cumulative_Actual_t'] = daily_sorted['Stoping_Actual_t'].cumsum()
        daily_sorted['Cumulative_Budget_t'] = daily_sorted['Stoping_Budget_t'].cumsum()
        daily_sorted['Cumulative_Actual_kg'] = daily_sorted['Stoping_Actual_kg'].cumsum()
        daily_sorted['Cumulative_Budget_kg'] = daily_sorted['Stoping_Budget_kg'].cumsum()
        cum_tonnes_sel = st.multiselect("Select Cumulative Tonnes", options=['Cumulative_Actual_t', 'Cumulative_Budget_t'], default=['Cumulative_Actual_t', 'Cumulative_Budget_t'], key="pd_cum_t")
        cum_gold_sel = st.multiselect("Select Cumulative Gold", options=['Cumulative_Actual_kg', 'Cumulative_Budget_kg'], default=['Cumulative_Actual_kg', 'Cumulative_Budget_kg'], key="pd_cum_g")
        cum_cols = st.columns(2)
        with cum_cols[0]:
            if cum_tonnes_sel:
                fig_cum_t = px.line(daily_sorted, x='Date', y=cum_tonnes_sel, labels={'value': 'Tonnes', 'variable': 'Metric'}, title='Cumulative Stoping Tonnes')
                fig_cum_t.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
                st.plotly_chart(fig_cum_t, use_container_width=True)
            else:
                st.warning("Select at least one cumulative tonnes series.")
        with cum_cols[1]:
            if cum_gold_sel:
                fig_cum_g = px.line(daily_sorted, x='Date', y=cum_gold_sel, labels={'value': 'Gold (kg)', 'variable': 'Metric'}, title='Cumulative Gold Production')
                fig_cum_g.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
                st.plotly_chart(fig_cum_g, use_container_width=True)
            else:
                st.warning("Select at least one cumulative gold series.")

    # Tab 5: Stope Drill‑Down
    with tabs[4]:
        st.markdown("#### Stope Drill‑Down")
        selected_stope = st.selectbox("Select a stope", options=sorted(df['Stope_ID'].dropna().unique()), key="pd_drill_stope")
        stope_data = df[df['Stope_ID'] == selected_stope].sort_values('Date')
        # Summary metrics
        cols = st.columns(3)
        with cols[0]:
            tot_t = stope_data['Stoping_Actual_t'].sum()
            bud_t = stope_data['Stoping_Budget_t'].sum()
            var_t = ((tot_t / bud_t) - 1) * 100 if bud_t else 0
            st.metric(f"Total Tonnes ({selected_stope})", f"{tot_t:,.0f} t", delta=f"{var_t:.1f}% vs Budget")
        with cols[1]:
            avg_g = stope_data['Stoping_Actual_gpt'].mean()
            bud_g = stope_data['Stoping_Budget_gpt'].mean()
            var_g = ((avg_g / bud_g) - 1) * 100 if bud_g else 0
            st.metric(f"Average Grade ({selected_stope})", f"{avg_g:.2f} g/t", delta=f"{var_g:.1f}% vs Budget")
        with cols[2]:
            tot_k = stope_data['Stoping_Actual_kg'].sum()
            bud_k = stope_data['Stoping_Budget_kg'].sum()
            var_k = ((tot_k / bud_k) - 1) * 100 if bud_k else 0
            st.metric(f"Total Gold ({selected_stope})", f"{tot_k:.2f} kg", delta=f"{var_k:.1f}% vs Budget")
        # Detailed history
        hist_cols = ['Date', 'BENCHES_Avg_Grade', 'Stoping_Actual_t', 'Stoping_Budget_t', 'Stoping_Actual_gpt', 'Stoping_Budget_gpt', 'Stoping_Actual_kg', 'Stoping_Budget_kg']
        history = stope_data[hist_cols].copy()
        history = history.rename(columns={
            'BENCHES_Avg_Grade': 'Bench Grade (g/t)', 'Stoping_Actual_t': 'Actual Tonnes', 'Stoping_Budget_t': 'Budget Tonnes',
            'Stoping_Actual_gpt': 'Actual Grade (g/t)', 'Stoping_Budget_gpt': 'Budget Grade (g/t)', 'Stoping_Actual_kg': 'Actual Gold (kg)', 'Stoping_Budget_kg': 'Budget Gold (kg)'
        })
        history['Tonnes Variance %'] = ((history['Actual Tonnes'] / history['Budget Tonnes']) - 1) * 100
        history['Grade Variance %'] = ((history['Actual Grade (g/t)'] / history['Budget Grade (g/t)']) - 1) * 100
        history['Gold Variance %'] = ((history['Actual Gold (kg)'] / history['Budget Gold (kg)']) - 1) * 100
        st.dataframe(history.style.format({
            'Date': lambda x: x.strftime('%Y-%m-%d'),
            'Bench Grade (g/t)': '{:.2f}', 'Actual Tonnes': '{:,.0f}', 'Budget Tonnes': '{:,.0f}', 'Tonnes Variance %': '{:+.1f}%',
            'Actual Grade (g/t)': '{:.2f}', 'Budget Grade (g/t)': '{:.2f}', 'Grade Variance %': '{:+.1f}%',
            'Actual Gold (kg)': '{:.2f}', 'Budget Gold (kg)': '{:.2f}', 'Gold Variance %': '{:+.1f}%'
        }), height=300)
        # Trends for selected stope
        sel_tonnes = st.multiselect("Select Tonnes Metrics", options=['Stoping_Actual_t', 'Stoping_Budget_t'], default=['Stoping_Actual_t', 'Stoping_Budget_t'], key="pd_drill_tonnes")
        if sel_tonnes:
            fig_dt = px.line(stope_data, x='Date', y=sel_tonnes, labels={'value': 'Tonnes', 'variable': 'Metric'}, title=f'Tonnes Trend ({selected_stope})')
            fig_dt.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_dt, use_container_width=True)
        sel_grade = st.multiselect("Select Grade Metrics", options=['Stoping_Actual_gpt', 'Stoping_Budget_gpt', 'BENCHES_Avg_Grade'], default=['Stoping_Actual_gpt', 'Stoping_Budget_gpt', 'BENCHES_Avg_Grade'], key="pd_drill_grade")
        if sel_grade:
            fig_dg = px.line(stope_data, x='Date', y=sel_grade, labels={'value': 'Grade (g/t)', 'variable': 'Metric'}, title=f'Grade Trend ({selected_stope})')
            fig_dg.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_dg, use_container_width=True)

    # Tab 6: Stoping vs Tramming Comparison
    with tabs[5]:
        st.markdown("#### Stoping vs Tramming Comparison")
        cols = st.columns(3)
        with cols[0]:
            stoping_t = filtered_df['Stoping_Actual_t'].sum()
            tramming_t = filtered_df['Tramming_Actual_t'].sum()
            diff_t = ((tramming_t / stoping_t) - 1) * 100 if stoping_t else 0
            st.metric("Total Stoping Tonnes", f"{stoping_t:,.0f} t")
            st.metric("Total Tramming Tonnes", f"{tramming_t:,.0f} t", delta=f"{diff_t:.1f}% vs Stoping")
        with cols[1]:
            stoping_g = filtered_df['Stoping_Actual_gpt'].mean()
            tramming_g = filtered_df['Tramming_Actual_gpt'].mean()
            diff_g = ((tramming_g / stoping_g) - 1) * 100 if stoping_g else 0
            st.metric("Average Stoping Grade", f"{stoping_g:.2f} g/t")
            st.metric("Average Tramming Grade", f"{tramming_g:.2f} g/t", delta=f"{diff_g:.1f}% vs Stoping")
        with cols[2]:
            stoping_k = filtered_df['Stoping_Actual_kg'].sum()
            tramming_k = filtered_df['Tramming_Actual_kg'].sum()
            diff_k = ((tramming_k / stoping_k) - 1) * 100 if stoping_k else 0
            st.metric("Total Stoping Gold", f"{stoping_k:.2f} kg")
            st.metric("Total Tramming Gold", f"{tramming_k:.2f} kg", delta=f"{diff_k:.1f}% vs Stoping")
        # Daily comparison charts
        st.markdown("##### Daily Comparison")
        comp = filtered_df.groupby('Date').agg({
            'Stoping_Actual_t': 'sum', 'Tramming_Actual_t': 'sum', 'Stoping_Budget_t': 'sum', 'Tramming_Budget_t': 'sum',
            'Stoping_Actual_gpt': 'mean', 'Tramming_Actual_gpt': 'mean', 'BENCHES_Avg_Grade': 'mean',
            'Stoping_Budget_gpt': 'mean', 'Tramming_Budget_gpt': 'mean',
            'Stoping_Actual_kg': 'sum', 'Tramming_Actual_kg': 'sum', 'Stoping_Budget_kg': 'sum', 'Tramming_Budget_kg': 'sum'
        }).reset_index()
        cc = st.columns(2)
        # Tonnes comparison
        tonnes_options = ['Stoping_Actual_t', 'Tramming_Actual_t', 'Stoping_Budget_t', 'Tramming_Budget_t']
        sel_tc = st.multiselect("Select Tonnes Series", options=tonnes_options, default=tonnes_options, key="pd_svt_tonnes")
        if sel_tc:
            fig_tc = px.line(comp, x='Date', y=sel_tc, labels={'value': 'Tonnes', 'variable': 'Series'}, title='Daily Tonnes: Stoping vs Tramming vs Budget')
            fig_tc.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            cc[0].plotly_chart(fig_tc, use_container_width=True)
        # Gold comparison
        gold_options = ['Stoping_Actual_kg', 'Tramming_Actual_kg', 'Stoping_Budget_kg', 'Tramming_Budget_kg']
        sel_gc = st.multiselect("Select Gold Series", options=gold_options, default=gold_options, key="pd_svt_gold")
        if sel_gc:
            fig_gc = px.line(comp, x='Date', y=sel_gc, labels={'value': 'Gold (kg)', 'variable': 'Series'}, title='Daily Gold Production: Stoping vs Tramming vs Budget')
            fig_gc.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            cc[0].plotly_chart(fig_gc, use_container_width=True)
        # Grade comparison
        grade_options = ['BENCHES_Avg_Grade', 'Stoping_Actual_gpt', 'Tramming_Actual_gpt', 'Stoping_Budget_gpt', 'Tramming_Budget_gpt']
        sel_grc = st.multiselect("Select Grade Series", options=grade_options, default=grade_options, key="pd_svt_grade")
        if sel_grc:
            fig_gc2 = px.line(comp, x='Date', y=sel_grc, labels={'value': 'Grade (g/t)', 'variable': 'Series'}, title='Daily Grade Comparison')
            fig_gc2.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            cc[1].plotly_chart(fig_gc2, use_container_width=True)


# ---------------------------------------------------------------------------
# Daily Report Update page
# ---------------------------------------------------------------------------
def run_daily_report_update_page():
    """Allow the user to generate or update geology work plan workbooks from
    uploaded daily report data.  The page accepts an Excel file with
    daily production data, derives the month automatically, copies a
    template workbook, runs the update script and returns the updated
    workbook to the user.  Alternatively the user may specify an
    explicit output filename.
    """
    st.markdown("## 📝 Daily Report Update")
    st.markdown(
        "Upload a monthly daily report workbook (.xlsx) to generate or update a geology work plan. "
        "If you do not provide an output filename, the system will create one automatically based on the month of the report."
    )

    if udr is None:
        st.error("The update script could not be loaded. Please ensure the v4 folder is present.")
        return

    uploaded_src = st.file_uploader("Upload Daily Report (source)", type=['xlsx', 'xls'], key="dr_src")
    dest_name_input = st.text_input("Output File Name (optional)", key="dr_dest")
    process_btn = st.button("Process Daily Report", key="dr_run")

    if not process_btn:
        return
    if uploaded_src is None:
        st.error("Please upload a daily report workbook before processing.")
        return

    # Save the uploaded source file with its original name so that the update
    # script can detect the month and year from the filename.  Use a
    # temporary directory to avoid collisions.
    src_basename = os.path.basename(uploaded_src.name)
    tmp_src_dir = tempfile.mkdtemp()
    src_path = os.path.join(tmp_src_dir, src_basename)
    with open(src_path, 'wb') as f:
        f.write(uploaded_src.getvalue())

    # Determine the output filename either from user input or from the
    # uploaded filename
    if dest_name_input:
        output_filename = dest_name_input if dest_name_input.lower().endswith('.xlsx') else dest_name_input + '.xlsx'
    else:
        output_filename = derive_output_filename(src_basename)

    # Copy the template workbook into a temporary location.  This file will
    # act as the destination workbook.  If the user specified a
    # filename, use it to name the final downloaded file; the underlying
    # working copy can still reside in a temporary directory.
    try:
        template_copy_path = copy_template_workbook()
    except Exception as e:
        st.error(f"Failed to locate template workbook: {e}")
        return

    # Destination path for the working copy
    dest_path = os.path.join(tempfile.gettempdir(), f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{output_filename}")
    try:
        shutil.copy(template_copy_path, dest_path)
    except Exception as e:
        st.error(f"Failed to prepare destination workbook: {e}")
        return

    # Set the update script's SRC_FILE and DEST_FILE variables to our paths
    try:
        # Reload the module so that any previous modifications to globals are
        # cleared.  This is important when multiple runs occur within the same
        # Streamlit session.
        import importlib
        importlib.reload(udr)
        # Overwrite globals
        udr.SRC_FILE = src_path
        udr.DEST_FILE = dest_path
        # Run the update
        with st.spinner("Running update script. Please wait..."):
            udr.main()
        # After running, read the updated file into memory
        with open(dest_path, 'rb') as f:
            updated_bytes = f.read()
        st.success("Daily report processed successfully!")
        st.download_button(
            label="Download Updated Work Plan",
            data=updated_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dr_download",
        )
        st.caption(f"File size: {len(updated_bytes):,} bytes")
    except Exception as e:
        st.error(f"An error occurred while running the update: {e}")


# ---------------------------------------------------------------------------
# Monthly Stope Performance page
# ---------------------------------------------------------------------------
def run_monthly_stope_performance_page():
    """Integrate the Monthly Stope Performance updater with consistent UI.

    Upload MSP workbook and optional supporting files, parse forecasts/actuals,
    update SUMMARY and PNM/MNP sheets, and provide a download of the updated file.
    """
    st.markdown("## 📈 Monthly Stope Performance Updater")
    st.markdown(
        "Upload the MSP workbook (.xlsx) and optional supporting files: 3‑Month Rolling forecasts, "
        "Actual Physicals, Mining & Milling Plan (for planned stopes), and Tramming Daily Reports (for PNM/MNP). "
        "Optionally include the August Daily Report to populate August actuals."
    )

    msp_file = st.file_uploader("Monthly Stope Performance workbook (required)", type=["xlsx"], key="msp_main")
    three_month_files = st.file_uploader("3‑Month Rolling files (optional, multiple)", type=["xlsx"], accept_multiple_files=True, key="msp_three")
    actual_files = st.file_uploader("Actual Physicals files (optional, multiple)", type=["xlsx"], accept_multiple_files=True, key="msp_actual")
    plan_files = st.file_uploader("Mining & Milling Plan files (optional, multiple)", type=["xlsx"], accept_multiple_files=True, key="msp_plan")
    tramming_files = st.file_uploader("Tramming Daily Report files (optional, multiple)", type=["xlsx"], accept_multiple_files=True, key="msp_tramming")
    daily_file = st.file_uploader("August Daily Report (optional)", type=["xlsx"], key="msp_daily")

    if st.button("Update Workbook", key="msp_update_btn"):
        if msp_file is None:
            st.error("Please upload the Monthly Stope Performance workbook.")
            return

        forecasts: Dict[int, Tuple[float, float, float]] = {}
        for uploaded in (three_month_files or []):
            parsed = parse_three_month_rolling(BytesIO(uploaded.getvalue()), uploaded.name)
            if parsed:
                for month_idx, triple in parsed.items():
                    forecasts[month_idx] = triple

        actuals: Dict[int, Tuple[float, float, float]] = {}
        for uploaded in (actual_files or []):
            parsed = parse_actual_physical(BytesIO(uploaded.getvalue()))
            if parsed:
                tonnes, grade, gold, month_idx = parsed
                actuals[month_idx] = (tonnes, grade, gold)

        if daily_file is not None:
            parsed = parse_august_daily_report(BytesIO(daily_file.getvalue()))
            if parsed:
                tonnes, grade, gold = parsed
                actuals[8] = (tonnes, grade, gold)

        planned_ids: Set[str] = set()
        for uploaded in (plan_files or []):
            ids = parse_underground_breaking_plan(BytesIO(uploaded.getvalue()))
            planned_ids.update(ids)

        pnm_by_month: Dict[int, Dict[str, Tuple[float, float, float]]] = {}
        mnp_by_month: Dict[int, Dict[str, Tuple[float, float, float]]] = {}
        for uploaded in (tramming_files or []):
            res = parse_tramming_detail(BytesIO(uploaded.getvalue()))
            if res:
                month_idx, tramming_data = res
                pnm_data: Dict[str, Tuple[float, float, float]] = {}
                mnp_data: Dict[str, Tuple[float, float, float]] = {}
                for stope_id, (b_t, a_t, b_g, a_g, b_au, a_au) in tramming_data.items():
                    diff_t = (b_t or 0) - (a_t or 0)
                    diff_g = (b_g or 0) - (a_g or 0)
                    diff_au = (b_au or 0) - (a_au or 0)
                    if diff_t != 0 or diff_g != 0 or diff_au != 0:
                        pnm_data[stope_id] = (diff_t, diff_g, diff_au)
                    if stope_id not in planned_ids:
                        mnp_data[stope_id] = (a_t or 0, a_g or 0, a_au or 0)
                if month_idx in pnm_by_month:
                    pnm_by_month[month_idx].update(pnm_data)
                else:
                    pnm_by_month[month_idx] = pnm_data
                if month_idx in mnp_by_month:
                    mnp_by_month[month_idx].update(mnp_data)
                else:
                    mnp_by_month[month_idx] = mnp_data

        try:
            updated_bytes = update_msp_workbook(BytesIO(msp_file.getvalue()), forecasts, actuals)
        except Exception as exc:
            st.error(f"Failed to update SUMMARY: {exc}")
            return

        try:
            wb = openpyxl.load_workbook(updated_bytes)
        except Exception as exc:
            st.error(f"Failed to open updated workbook for PNM/MNP updates: {exc}")
            return

        for month_idx, pnm_data in pnm_by_month.items():
            mnp_data = mnp_by_month.get(month_idx, {})
            update_pnm_mnp_sheet(wb, pnm_data, mnp_data, month_idx)

        final_bytes = BytesIO()
        wb.save(final_bytes)
        final_bytes.seek(0)
        out_name = f"Monthly_Stope_Performance_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.success("Workbook updated successfully!")
        st.download_button(
            label="Download Updated MSP Workbook",
            data=final_bytes.getvalue(),
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="msp_download",
        )
        st.caption(f"File size: {final_bytes.getbuffer().nbytes:,} bytes | Forecast months: {len(forecasts)} | Actual months: {len(actuals)} | PNM months: {len(pnm_by_month)} | MNP months: {len(mnp_by_month)}")

# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def main():
    """Dispatch to the selected page.  The sidebar provides a simple
    navigation menu.  Each page is encapsulated in its own function.
    """
    st.sidebar.title("Navigation")
    page = st.sidebar.radio(
        "Choose a page:",
        [
            "Mining Data Processor",
            "Production Dashboard",
            "Daily Report Update",
            "Monthly Stope Performance",
        ],
    )
    if page == "Mining Data Processor":
        run_mining_processor_page()
    elif page == "Production Dashboard":
        run_production_dashboard_page()
    elif page == "Daily Report Update":
        run_daily_report_update_page()
    elif page == "Monthly Stope Performance":
        run_monthly_stope_performance_page()


if __name__ == '__main__':
    main()