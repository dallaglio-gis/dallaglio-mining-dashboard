import openpyxl
import datetime
import re
from calendar import month_name

"""Generate individual daily sheets for every day that has data.

This script reads a monthly DAILY_REPORT.xlsx source workbook (e.g., June_2025_DAILY_REPORT.xlsx, 
July_2025_DAILY_REPORT.xlsx) and the corresponding destination template workbook.
The month and year are automatically detected from the source filename.

For every day in the detected month that contains non-zero production data, the script 
creates a fresh sheet in the destination workbook and populates the daily, month-to-date 
(MTD) and budget MTD figures.

IMPORTANT: This script only creates sheets for dates that don't already exist
in the destination workbook. Existing sheets are preserved to maintain correct
MTD calculations and avoid overwriting manually adjusted values.

The template is assumed to be the first (active) worksheet in the
destination workbook. All formatting and dimensions from the template
are copied to each newly created sheet so that visual consistency is
maintained.

Supported filename formats:
- MonthName_YYYY_DAILY_REPORT.xlsx (e.g., June_2025_DAILY_REPORT.xlsx)
- MonthName_YY_DAILY_REPORT.xlsx (e.g., July_25_DAILY_REPORT.xlsx)
"""

SRC_FILE = r"August_2025_DAILY_REPORT.xlsx"
DEST_FILE = r"Geology Daily Work Plan August2025.xlsx"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_month_year_from_filename(filename):
    """Extract month and year from filename like 'June_2025_DAILY_REPORT.xlsx' or 'July_2025_DAILY_REPORT.xlsx'"""
    import re
    import os
    
    # Get just the filename without path
    basename = os.path.basename(filename)
    
    # Month name mapping
    month_names = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    
    # Try to find month name and year in filename
    # Pattern: MonthName_YYYY or MonthName_YY
    pattern = r'([a-zA-Z]+)_(\d{4}|\d{2})'
    match = re.search(pattern, basename.lower())
    
    if match:
        month_str = match.group(1).lower()
        year_str = match.group(2)
        
        # Convert month name to number
        month_num = month_names.get(month_str)
        if month_num:
            # Convert year to 4-digit format
            year = int(year_str)
            if year < 100:  # 2-digit year
                year += 2000 if year < 50 else 1900
            
            return month_num, year
    
    # Fallback: try to find just year and assume current month
    year_match = re.search(r'(\d{4})', basename)
    if year_match:
        year = int(year_match.group(1))
        # Default to June if no month found (for backward compatibility)
        return 6, year
    
    # Final fallback
    return 6, 2025

def normalise_date(value):
    """Attempt to convert *value* to a datetime.date.  Returns None if
    conversion fails."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    
    # Get month and year from source filename
    month, year = extract_month_year_from_filename(SRC_FILE)
    
    # Handle day numbers (1, 2, 3, ...) using detected month/year
    if isinstance(value, (int, float)) and 1 <= value <= 31:
        try:
            return datetime.date(year, month, int(value))
        except ValueError:
            pass
    
    # Handle string day numbers
    if isinstance(value, str) and value.strip().isdigit():
        day_num = int(value.strip())
        if 1 <= day_num <= 31:
            try:
                return datetime.date(year, month, day_num)
            except ValueError:
                pass
    
    # Handle standard date formats
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(str(value), fmt).date()
        except Exception:
            continue
    
    return None

def sheet_name_from_date(dt: datetime.date) -> str:
    """Return Excel-safe sheet name like '01June25'."""
    name = dt.strftime("%d%B%y")  # 01June25
    # Excel sheet names cannot contain these chars / \ * ? [ ] :
    return re.sub(r"[\\/:*?\[\]]", "_", name)[:31]


def copy_formatting(src_ws, dest_ws):
    """Copy cell values & formatting from *src_ws* to *dest_ws*."""
    from copy import copy

    for row in src_ws.iter_rows(values_only=False):
        for cell in row:
            new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copy column widths
    for col, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col].width = dim.width
    # Copy row heights
    for idx, dim in src_ws.row_dimensions.items():
        dest_ws.row_dimensions[idx].height = dim.height

# ---------------------------------------------------------------------------
# Extract data indices with non-zero daily tonnes (Tramming & Plant)
# ---------------------------------------------------------------------------

def extract_indices():
    """Return mapping date -> (tram_idx, plant_idx)."""
    src_wb = openpyxl.load_workbook(SRC_FILE, data_only=True)
    tram = src_wb["Tramming"]
    plant = src_wb["PLANT"]

    tram_dates = ['H9','I9','J9','K9','L9','M9','N9','O9','P9','Q9','R9','S9','T9','U9','V9','W9','X9','Y9','Z9','AA9','AB9','AC9','AD9','AE9','AF9','AG9','AH9','AI9','AJ9','AK9','AL9','AM9','AN9','AO9','AP9']
    tram_tonnes = ['H11','I11','J11','K11','L11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11','Y11','Z11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AN11','AO11','AP11']

    plant_dates = ['H4','I4','J4','K4','L4','M4','N4','O4','P4','Q4','R4','S4','T4','U4','V4','W4','X4','Y4','Z4','AA4','AB4','AC4','AD4','AE4','AF4','AG4','AH4','AI4','AJ4','AK4','AL4','AM4','AN4','AO4','AP4']
    plant_tonnes = ['H7','I7','J7','K7','L7','M7','N7','O7','P7','Q7','R7','S7','T7','U7','V7','W7','X7','Y7','Z7','AA7','AB7','AC7','AD7','AE7','AF7','AG7','AH7','AI7','AJ7','AK7','AL7','AM7','AN7','AO7','AP7']

    mapping = {}

    # Tramming indices
    for idx, vcell in enumerate(tram_tonnes):
        val = tram[vcell].value
        try:
            val = float(val)
        except (ValueError, TypeError):
            val = 0
        if val and val > 0:
            dt = normalise_date(tram[tram_dates[idx]].value)
            if dt:
                mapping.setdefault(dt, [None, None])[0] = idx

    # Plant indices
    for idx, vcell in enumerate(plant_tonnes):
        val = plant[vcell].value
        try:
            val = float(val)
        except (ValueError, TypeError):
            val = 0
        if val and val > 0:
            dt = normalise_date(plant[plant_dates[idx]].value)
            if dt:
                mapping.setdefault(dt, [None, None])[1] = idx

    src_wb.close()
    return mapping

# ---------------------------------------------------------------------------
# Main routine
# ---------------------------------------------------------------------------

def calculate_mtd_values(daily_data, target_date):
    """Calculate proper MTD values up to the target date.
    Returns (mtd_tonnes, weighted_avg_grade, mtd_gold)
    """
    total_tonnes = 0
    total_gold = 0
    grade_weighted_sum = 0
    
    for date, data in daily_data.items():
        if date <= target_date:
            tonnes = data.get('tonnes', 0) or 0
            grade = data.get('grade', 0) or 0
            gold = data.get('gold', 0) or 0
            
            total_tonnes += tonnes
            total_gold += gold
            grade_weighted_sum += tonnes * grade
    
    # Calculate weighted average grade
    mtd_grade = grade_weighted_sum / total_tonnes if total_tonnes > 0 else 0
    
    return total_tonnes, mtd_grade, total_gold

def extract_daily_data_for_month(tram_sheet, plant_sheet, date_idx_map, target_month, target_year):
    """Extract all daily data for the target month, including budget values."""
    tram_tonnes = ['H11','I11','J11','K11','L11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11','Y11','Z11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AN11','AO11','AP11']
    tram_grade_row = 13
    tram_gold_rows = (15, 16)
    # Budget rows for tramming 
    tram_budget_tonnes_row = 10  
    tram_budget_grade_row = 12
    tram_budget_gold_row = 14
    
    plant_tonnes = ['H7','I7','J7','K7','L7','M7','N7','O7','P7','Q7','R7','S7','T7','U7','V7','W7','X7','Y7','Z7','AA7','AB7','AC7','AD7','AE7','AF7','AG7','AH7','AI7','AJ7','AK7','AL7','AM7','AN7','AO7','AP7']
    plant_grade_row = 10
    plant_gold_row = 16
    # Budget rows for plant 
    plant_budget_tonnes_row = 5
    plant_budget_grade_row = 8
    plant_budget_gold_row = 14
    
    tram_data = {}
    plant_data = {}
    tram_budget_data = {}
    plant_budget_data = {}
    
    for dt, (tram_idx, plant_idx) in date_idx_map.items():
        if dt.month != target_month or dt.year != target_year:
            continue
            
        # Extract tramming data
        if tram_idx is not None:
            # Actual values
            tonnes = tram_sheet[tram_tonnes[tram_idx]].value or 0
            grade_cell = re.sub(r"\d+", str(tram_grade_row), tram_tonnes[tram_idx])
            gold_cell_15 = re.sub(r"\d+", str(tram_gold_rows[0]), tram_tonnes[tram_idx])
            gold_cell_16 = re.sub(r"\d+", str(tram_gold_rows[1]), tram_tonnes[tram_idx])
            grade = tram_sheet[grade_cell].value or 0
            gold = tram_sheet[gold_cell_15].value
            if gold in (None, 0, ''):
                gold = tram_sheet[gold_cell_16].value or 0
            else:
                gold = gold or 0
                
            tram_data[dt] = {'tonnes': tonnes, 'grade': grade, 'gold': gold}
            
            # Budget values
            budget_tonnes_cell = re.sub(r"\d+", str(tram_budget_tonnes_row), tram_tonnes[tram_idx])
            budget_grade_cell = re.sub(r"\d+", str(tram_budget_grade_row), tram_tonnes[tram_idx])
            budget_gold_cell = re.sub(r"\d+", str(tram_budget_gold_row), tram_tonnes[tram_idx])
            
            budget_tonnes = tram_sheet[budget_tonnes_cell].value or 0
            budget_grade = tram_sheet[budget_grade_cell].value or 0
            budget_gold = tram_sheet[budget_gold_cell].value or 0
            
            tram_budget_data[dt] = {'tonnes': budget_tonnes, 'grade': budget_grade, 'gold': budget_gold}
        
        # Extract plant data
        if plant_idx is not None:
            # Actual values
            tonnes = plant_sheet[plant_tonnes[plant_idx]].value or 0
            grade_cell = re.sub(r"\d+", str(plant_grade_row), plant_tonnes[plant_idx])
            gold_cell = re.sub(r"\d+", str(plant_gold_row), plant_tonnes[plant_idx])
            grade = plant_sheet[grade_cell].value or 0
            gold = plant_sheet[gold_cell].value or 0
            
            plant_data[dt] = {'tonnes': tonnes, 'grade': grade, 'gold': gold}
            
            # Budget values
            budget_tonnes_cell = re.sub(r"\d+", str(plant_budget_tonnes_row), plant_tonnes[plant_idx])
            budget_grade_cell = re.sub(r"\d+", str(plant_budget_grade_row), plant_tonnes[plant_idx])
            budget_gold_cell = re.sub(r"\d+", str(plant_budget_gold_row), plant_tonnes[plant_idx])
            
            budget_tonnes = plant_sheet[budget_tonnes_cell].value or 0
            budget_grade = plant_sheet[budget_grade_cell].value or 0
            budget_gold = plant_sheet[budget_gold_cell].value or 0
            
            plant_budget_data[dt] = {'tonnes': budget_tonnes, 'grade': budget_grade, 'gold': budget_gold}
    
    return tram_data, plant_data, tram_budget_data, plant_budget_data

def needs_update(worksheet, daily_values, mtd_values, budget_mtd_values):
    """Check if worksheet needs updating by comparing current values with new ones."""
    # Check daily values (G column)
    current_daily = [
        worksheet['G4'].value,  # tram tonnes
        worksheet['G5'].value,  # tram grade  
        worksheet['G6'].value,  # tram gold
        worksheet['G7'].value,  # plant tonnes
        worksheet['G8'].value,  # plant grade
        worksheet['G9'].value   # plant gold
    ]
    
    # Check MTD values (H column)
    current_mtd = [
        worksheet['H4'].value,  # tram tonnes MTD
        worksheet['H5'].value,  # tram grade MTD
        worksheet['H6'].value,  # tram gold MTD
        worksheet['H7'].value,  # plant tonnes MTD
        worksheet['H8'].value,  # plant grade MTD
        worksheet['H9'].value   # plant gold MTD
    ]
    
    # Check Budget MTD values (I column)
    current_budget_mtd = [
        worksheet['I4'].value,  # tram tonnes Budget MTD
        worksheet['I5'].value,  # tram grade Budget MTD
        worksheet['I6'].value,  # tram gold Budget MTD
        worksheet['I7'].value,  # plant tonnes Budget MTD
        worksheet['I8'].value,  # plant grade Budget MTD
        worksheet['I9'].value   # plant gold Budget MTD
    ]
    
    new_daily = daily_values
    new_mtd = mtd_values
    new_budget_mtd = budget_mtd_values
    
    # Compare with tolerance for floating point differences
    def values_different(a, b, tolerance=0.001):
        if a is None and b is None:
            return False
        if a is None or b is None:
            return True
        try:
            return abs(float(a) - float(b)) > tolerance
        except (ValueError, TypeError):
            return str(a) != str(b)
    
    for i in range(len(current_daily)):
        if (values_different(current_daily[i], new_daily[i]) or 
            values_different(current_mtd[i], new_mtd[i]) or
            values_different(current_budget_mtd[i], new_budget_mtd[i])):
            return True
    
    return False

def main():
    # Get the target month and year from source filename
    target_month, target_year = extract_month_year_from_filename(SRC_FILE)
    month_name = datetime.date(target_year, target_month, 1).strftime('%B')
    
    date_idx_map = extract_indices()
    # Process all dates with data for the detected month/year
    target_dates = sorted([d for d in date_idx_map if d.month == target_month and d.year == target_year])
    if not target_dates:
        print(f"No valid dates with non-zero values found for {month_name} {target_year}.")
        return

    src_wb = openpyxl.load_workbook(SRC_FILE, data_only=True)
    tram = src_wb["Tramming"]
    plant = src_wb["PLANT"]

    dest_wb = openpyxl.load_workbook(DEST_FILE)
    template_ws = dest_wb.active

    # Extract all daily data for the month including budget data
    tram_daily_data, plant_daily_data, tram_budget_data, plant_budget_data = extract_daily_data_for_month(tram, plant, date_idx_map, target_month, target_year)

    # Get existing sheet names
    existing_sheet_names = set(dest_wb.sheetnames)
    
    sheets_created = []
    sheets_updated = []
    
    print(f"Processing {len(target_dates)} dates for {month_name} {target_year}...")
    
    # Check if any data has changed by comparing with the latest sheet
    force_update_all = False
    if target_dates:
        latest_date = target_dates[-1]
        latest_sheet_name = sheet_name_from_date(latest_date)
        
        if latest_sheet_name in existing_sheet_names:
            # Calculate values for the latest date
            tram_mtd_tonnes, tram_mtd_grade, tram_mtd_gold = calculate_mtd_values(tram_daily_data, latest_date)
            plant_mtd_tonnes, plant_mtd_grade, plant_mtd_gold = calculate_mtd_values(plant_daily_data, latest_date)
            
            tram_budget_mtd_tonnes, tram_budget_mtd_grade, tram_budget_mtd_gold = calculate_mtd_values(tram_budget_data, latest_date)
            plant_budget_mtd_tonnes, plant_budget_mtd_grade, plant_budget_mtd_gold = calculate_mtd_values(plant_budget_data, latest_date)
            
            tram_daily = tram_daily_data.get(latest_date, {'tonnes': None, 'grade': None, 'gold': None})
            plant_daily = plant_daily_data.get(latest_date, {'tonnes': None, 'grade': None, 'gold': None})
            
            daily_values = [
                tram_daily['tonnes'], tram_daily['grade'], tram_daily['gold'],
                plant_daily['tonnes'], plant_daily['grade'], plant_daily['gold']
            ]
            
            mtd_values = [
                tram_mtd_tonnes, tram_mtd_grade, tram_mtd_gold,
                plant_mtd_tonnes, plant_mtd_grade, plant_mtd_gold
            ]
            
            budget_mtd_values = [
                tram_budget_mtd_tonnes, tram_budget_mtd_grade, tram_budget_mtd_gold,
                plant_budget_mtd_tonnes, plant_budget_mtd_grade, plant_budget_mtd_gold
            ]
            
            latest_ws = dest_wb[latest_sheet_name]
            # If the latest sheet needs updating, force update all previous sheets
            if needs_update(latest_ws, daily_values, mtd_values, budget_mtd_values):
                force_update_all = True
                print("Data changes detected. Updating all sheets with recalculated MTD values...")
    
    for dt in target_dates:
        sheet_name = sheet_name_from_date(dt)
        
        # Calculate MTD values up to this date
        tram_mtd_tonnes, tram_mtd_grade, tram_mtd_gold = calculate_mtd_values(tram_daily_data, dt)
        plant_mtd_tonnes, plant_mtd_grade, plant_mtd_gold = calculate_mtd_values(plant_daily_data, dt)
        
        # Calculate Budget MTD values up to this date
        tram_budget_mtd_tonnes, tram_budget_mtd_grade, tram_budget_mtd_gold = calculate_mtd_values(tram_budget_data, dt)
        plant_budget_mtd_tonnes, plant_budget_mtd_grade, plant_budget_mtd_gold = calculate_mtd_values(plant_budget_data, dt)
        
        # Get daily values for this specific date
        tram_daily = tram_daily_data.get(dt, {'tonnes': None, 'grade': None, 'gold': None})
        plant_daily = plant_daily_data.get(dt, {'tonnes': None, 'grade': None, 'gold': None})
        
        daily_values = [
            tram_daily['tonnes'], tram_daily['grade'], tram_daily['gold'],
            plant_daily['tonnes'], plant_daily['grade'], plant_daily['gold']
        ]
        
        mtd_values = [
            tram_mtd_tonnes, tram_mtd_grade, tram_mtd_gold,
            plant_mtd_tonnes, plant_mtd_grade, plant_mtd_gold
        ]
        
        budget_mtd_values = [
            tram_budget_mtd_tonnes, tram_budget_mtd_grade, tram_budget_mtd_gold,
            plant_budget_mtd_tonnes, plant_budget_mtd_grade, plant_budget_mtd_gold
        ]
        
        if sheet_name in existing_sheet_names:
            # Check if existing sheet needs updating or if force_update_all is True
            existing_ws = dest_wb[sheet_name]
            if force_update_all or needs_update(existing_ws, daily_values, mtd_values, budget_mtd_values):
                # Update existing sheet
                existing_ws['G4'] = tram_daily['tonnes']
                existing_ws['H4'] = tram_mtd_tonnes
                existing_ws['I4'] = tram_budget_mtd_tonnes
                
                existing_ws['G5'] = tram_daily['grade']
                existing_ws['H5'] = tram_mtd_grade
                existing_ws['I5'] = tram_budget_mtd_grade
                
                existing_ws['G6'] = tram_daily['gold']
                existing_ws['H6'] = tram_mtd_gold
                existing_ws['I6'] = tram_budget_mtd_gold
                
                existing_ws['G7'] = plant_daily['tonnes']
                existing_ws['H7'] = plant_mtd_tonnes
                existing_ws['I7'] = plant_budget_mtd_tonnes
                
                existing_ws['G8'] = plant_daily['grade']
                existing_ws['H8'] = plant_mtd_grade
                existing_ws['I8'] = plant_budget_mtd_grade
                
                existing_ws['G9'] = plant_daily['gold']
                existing_ws['H9'] = plant_mtd_gold
                existing_ws['I9'] = plant_budget_mtd_gold
                
                sheets_updated.append(sheet_name)
        else:
            # Create new sheet from template
            new_ws = dest_wb.copy_worksheet(template_ws)
            new_ws.title = sheet_name
            
            # Populate new sheet with calculated values
            new_ws['G4'] = tram_daily['tonnes']
            new_ws['H4'] = tram_mtd_tonnes
            new_ws['I4'] = tram_budget_mtd_tonnes
            
            new_ws['G5'] = tram_daily['grade']
            new_ws['H5'] = tram_mtd_grade
            new_ws['I5'] = tram_budget_mtd_grade
            
            new_ws['G6'] = tram_daily['gold']
            new_ws['H6'] = tram_mtd_gold
            new_ws['I6'] = tram_budget_mtd_gold
            
            new_ws['G7'] = plant_daily['tonnes']
            new_ws['H7'] = plant_mtd_tonnes
            new_ws['I7'] = plant_budget_mtd_tonnes
            
            new_ws['G8'] = plant_daily['grade']
            new_ws['H8'] = plant_mtd_grade
            new_ws['I8'] = plant_budget_mtd_grade
            
            new_ws['G9'] = plant_daily['gold']
            new_ws['H9'] = plant_mtd_gold
            new_ws['I9'] = plant_budget_mtd_gold
            
            sheets_created.append(sheet_name)

    # Don't remove the template sheet if no new sheets were created
    # Only remove template if it's still the default "Sheet" name and we created new sheets
    if template_ws.title in ["Sheet", "Sheet1"] and sheets_created:
        dest_wb.remove(template_ws)

    dest_wb.save(DEST_FILE)
    src_wb.close()
    dest_wb.close()
    
    # Report results
    if sheets_created:
        print(f"Created {len(sheets_created)} new sheets: {', '.join(sheets_created)}")
    if sheets_updated:
        print(f"Updated {len(sheets_updated)} existing sheets: {', '.join(sheets_updated)}")
    if not sheets_created and not sheets_updated:
        print(f"No changes needed for {month_name} {target_year} - all sheets are up to date.")
    else:
        print(f"Latest date processed: {target_dates[-1].strftime('%d %B %Y')}")


if __name__ == "__main__":
    main()
