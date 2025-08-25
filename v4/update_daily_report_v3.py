import openpyxl
import datetime

# === UPDATE THESE FILE PATHS ===
src_file = r'June_2025_DAILY_REPORT.xlsx'
dest_file = r'Geology Daily Work Plan June2025.xlsx'

import re
from calendar import month_name
from copy import copy

def get_latest_nonzero_date(ws, value_cells, date_cells):
    """Return a tuple (date, index) where `date` represents the last date for
    which the corresponding value in `value_cells` is > 0.

    If no non-zero value is found, returns `(None, None)`.
    """
    latest_idx = None

    for i, vcell in enumerate(value_cells):
        value = ws[vcell].value
        try:
            value = float(value)
        except (ValueError, TypeError):
            value = 0
        if value and value > 0:
            latest_idx = i

    if latest_idx is None:
        return None, None

    date_value = ws[date_cells[latest_idx]].value

    # Normalise date_value to a datetime where possible
    if isinstance(date_value, datetime.datetime):
        return date_value, latest_idx

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(str(date_value), fmt), latest_idx
        except Exception:
            continue
    # As a last resort return raw string for date
    return date_value, latest_idx

def get_all_dates_with_data(ws, value_cells, date_cells):
    """Return list of (date, index) tuples for all dates with non-zero values."""
    dates_with_data = []
    
    for i, vcell in enumerate(value_cells):
        value = ws[vcell].value
        date_value = ws[date_cells[i]].value
        
        try:
            value = float(value)
        except (ValueError, TypeError):
            value = 0
            
        if value and value > 0:
            # Convert day number to full date for June 2025
            if isinstance(date_value, (int, float)) and 1 <= date_value <= 31:
                try:
                    full_date = datetime.datetime(2025, 6, int(date_value))
                    dates_with_data.append((full_date, i))
                    continue
                except ValueError:
                    pass
            
            # Handle other date formats
            if isinstance(date_value, datetime.datetime):
                dates_with_data.append((date_value, i))
            else:
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
                    try:
                        parsed_date = datetime.datetime.strptime(str(date_value), fmt)
                        dates_with_data.append((parsed_date, i))
                        break
                    except Exception:
                        continue
    
    return sorted(dates_with_data, key=lambda x: x[0])

def get_last_existing_sheet_date(dest_wb):
    """Find the latest date from existing sheet names in the workbook."""
    latest_date = None
    
    for sheet_name in dest_wb.sheetnames:
        # Try to parse sheet names like "22June25"
        match = re.match(r'(\d{1,2})([A-Za-z]+)(\d{2})', sheet_name)
        if match:
            day, month_name_str, year = match.groups()
            try:
                # Convert month name to number
                month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                              'July', 'August', 'September', 'October', 'November', 'December']
                month_num = None
                for i, m in enumerate(month_names):
                    if m.lower().startswith(month_name_str.lower()):
                        month_num = i + 1
                        break
                
                if month_num:
                    # Assume 20xx for year
                    full_year = 2000 + int(year)
                    sheet_date = datetime.datetime(full_year, month_num, int(day))
                    
                    if latest_date is None or sheet_date > latest_date:
                        latest_date = sheet_date
            except:
                continue
    
    return latest_date

def format_sheet_name(dt):
    """Format date as sheet name like '23June25'."""
    name = dt.strftime("%d%B%y")
    # Excel sheet names cannot contain these chars / \ * ? [ ] :
    return re.sub(r'[\\/:*?\[\]]', '_', name)[:31]

# --- 1. Open workbooks ---
src_wb = openpyxl.load_workbook(src_file, data_only=True)
dest_wb = openpyxl.load_workbook(dest_file)

# --- 2. Extract from Tramming ---
tram = src_wb['Tramming']

# Daily dates: for the latest nonzero logic
tram_dates = ['H9','I9','J9','K9','L9','M9','N9','O9','P9','Q9','R9','S9','T9','U9','V9','W9','X9','Y9','Z9','AA9','AB9', 'AC9','AD9','AE9','AF9','AG9','AH9','AI9','AJ9','AK9','AL9','AM9','AN9','AO9','AP9']
tram_tonnes = ['H11','I11','J11','K11','L11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11','Y11','Z11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AN11','AO11','AP11']

# MTD values
tram_mtd_tonnes = tram['D10'].value
tram_mtd_grade = tram['D11'].value
tram_mtd_gold = tram['D12'].value

# Budget MTD
tram_budget_tonnes = tram['C10'].value
tram_budget_grade = tram['C11'].value
tram_budget_gold = tram['C12'].value

# --- 3. Extract from PLANT ---
plant = src_wb['PLANT']

# Daily dates: for latest nonzero logic
plant_dates = ['H4','I4','J4','K4','L4','M4','N4','O4','P4','Q4','R4','S4','T4','U4','V4','W4','X4','Y4','Z4','AA4','AB4','AC4','AD4','AE4','AF4','AG4','AH4','AI4','AJ4','AK4','AL4','AM4','AN4','AO4','AP4']
plant_tonnes = ['H7','I7','J7','K7','L7','M7','N7','O7','P7','Q7','R7','S7','T7','U7','V7','W7','X7','Y7','Z7','AA7','AB7','AC7','AD7','AE7','AF7','AG7','AH7','AI7','AJ7','AK7','AL7','AM7','AN7','AO7','AP7']

# MTD values
plant_mtd_tonnes = plant['C13'].value
plant_mtd_grade = plant['D13'].value
plant_mtd_gold = plant['E13'].value

# Budget MTD
plant_budget_tonnes = plant['C10'].value
plant_budget_grade = plant['D10'].value
plant_budget_gold = plant['E10'].value

# --- 4. Get all dates with data ---
tram_dates_with_data = get_all_dates_with_data(tram, tram_tonnes, tram_dates)
plant_dates_with_data = get_all_dates_with_data(plant, plant_tonnes, plant_dates)

# Combine and get unique dates
all_dates_with_data = {}
for date, idx in tram_dates_with_data:
    all_dates_with_data[date] = {'tram_idx': idx, 'plant_idx': None}

for date, idx in plant_dates_with_data:
    if date in all_dates_with_data:
        all_dates_with_data[date]['plant_idx'] = idx
    else:
        all_dates_with_data[date] = {'tram_idx': None, 'plant_idx': idx}

# Sort dates
sorted_dates = sorted(all_dates_with_data.keys())

# --- 5. Find last existing sheet date ---
last_existing_date = get_last_existing_sheet_date(dest_wb)

# --- 6. Determine which dates need new sheets ---
dates_to_create = []
if last_existing_date:
    for date in sorted_dates:
        if date > last_existing_date:
            dates_to_create.append(date)
else:
    # If no existing sheets found, create for all dates
    dates_to_create = sorted_dates

# --- 7. Get template sheet ---
template_sheet = dest_wb.active

# --- 8. Create sheets for missing dates ---
sheets_created = []
for date in dates_to_create:
    indices = all_dates_with_data[date]
    tram_idx = indices['tram_idx']
    plant_idx = indices['plant_idx']
    
    # Calculate daily values
    tram_daily_tonnes = tram_daily_grade = tram_daily_gold = None
    plant_daily_tonnes = plant_daily_grade = plant_daily_gold = None
    
    if tram_idx is not None:
        tram_daily_tonnes = tram[tram_tonnes[tram_idx]].value
        grade_cell = re.sub(r'\d+', '13', tram_tonnes[tram_idx])
        gold_cell_15 = re.sub(r'\d+', '15', tram_tonnes[tram_idx])
        gold_cell_16 = re.sub(r'\d+', '16', tram_tonnes[tram_idx])
        tram_daily_grade = tram[grade_cell].value
        tram_daily_gold = tram[gold_cell_15].value if tram[gold_cell_15].value not in (None, 0, '') else tram[gold_cell_16].value
    
    if plant_idx is not None:
        plant_daily_tonnes = plant[plant_tonnes[plant_idx]].value
        plant_grade_cell = re.sub(r'\d+', '10', plant_tonnes[plant_idx])
        plant_gold_cell = re.sub(r'\d+', '16', plant_tonnes[plant_idx])
        plant_daily_grade = plant[plant_grade_cell].value
        plant_daily_gold = plant[plant_gold_cell].value
    
    # Create new sheet name
    sheet_name = format_sheet_name(date)
    
    # Remove existing sheet if it exists
    if sheet_name in dest_wb.sheetnames:
        dest_wb.remove(dest_wb[sheet_name])
    
    # Create new sheet at the end
    new_sheet = dest_wb.create_sheet(sheet_name)
    
    # Copy formatting and structure from template
    for row in template_sheet.iter_rows(values_only=False):
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    
    # Copy dimensions
    for col_key, col_dim in template_sheet.column_dimensions.items():
        new_sheet.column_dimensions[col_key].width = col_dim.width
    for row_key, row_dim in template_sheet.row_dimensions.items():
        new_sheet.row_dimensions[row_key].height = row_dim.height
    
    # Populate data cells
    # Tramming (Mining) (Tonnes, Grade, Gold): G4:H4:I4, G5:H5:I5, G6:H6:I6
    new_sheet['G4'] = tram_daily_tonnes
    new_sheet['H4'] = tram_mtd_tonnes
    new_sheet['I4'] = tram_budget_tonnes
    
    new_sheet['G5'] = tram_daily_grade
    new_sheet['H5'] = tram_mtd_grade
    new_sheet['I5'] = tram_budget_grade
    
    new_sheet['G6'] = tram_daily_gold
    new_sheet['H6'] = tram_mtd_gold
    new_sheet['I6'] = tram_budget_gold
    
    # Milling (PLANT) (Tonnes, Grade, Gold): G7:H7:I7, G8:H8:I8, G9:H9:I9
    new_sheet['G7'] = plant_daily_tonnes
    new_sheet['H7'] = plant_mtd_tonnes
    new_sheet['I7'] = plant_budget_tonnes
    
    new_sheet['G8'] = plant_daily_grade
    new_sheet['H8'] = plant_mtd_grade
    new_sheet['I8'] = plant_budget_grade
    
    new_sheet['G9'] = plant_daily_gold
    new_sheet['H9'] = plant_mtd_gold
    new_sheet['I9'] = plant_budget_gold
    
    sheets_created.append(sheet_name)

# --- 9. Save & cleanup ---
dest_wb.save(dest_file)
src_wb.close()
dest_wb.close()

# --- 10. Output message ---
if sheets_created:
    latest_date = max(dates_to_create)
    message_date = latest_date.strftime("%A %d %B %Y")
    with open("update_message.txt", "w", encoding='utf-8') as f:
        f.write(f"Created {len(sheets_created)} missing daily reports. Latest: {message_date}")
    
    print(f"Completed: Created {len(sheets_created)} missing daily reports")
    print(f"Sheets created: {', '.join(sheets_created)}")
    print(f"Latest records updated for: {message_date}")
else:
    with open("update_message.txt", "w", encoding='utf-8') as f:
        f.write("No missing daily reports found - all sheets are up to date")
    print("No missing daily reports found - all sheets are up to date")
