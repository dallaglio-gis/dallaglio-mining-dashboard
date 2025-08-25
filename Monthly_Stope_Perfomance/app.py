"""
Streamlit application to automate updating of the Monthly Stope Performance
workbook. Users can upload the primary workbook alongside a collection of
supporting spreadsheets (3‑month rolling forecasts, monthly actual physical
reports and an August daily report).  The tool parses the required metrics from
each source and overwrites the SHORT‑TERM FORECAST and ACTUAL sections in
the summary sheet of the Monthly Stope Performance file.

Parsing logic
-------------

* **3‑Month Rolling files** –  These workbooks typically contain a sheet
  named ``MINE PLAN`` where milling projections are tabulated.  The rows
  labelled “MILLING” and “Tonnes (t)” contain projected tonnages for the
  three months covered by the file.  Immediately below resides a row
  labelled “Grade - (g/t)”.  A separate row labelled “3 Month Rolling ”
  (note the trailing space) holds projected gold (kg).  The first numeric
  value in each of these rows corresponds to the month in the filename
  (e.g. ``3 Months Rolling_June 2025.xlsx`` → June).  Should this search
  fail, the parsing function returns ``None`` for that metric.

* **Actual Physicals files** –  For each monthly file the ``Dashboard``
  sheet contains a “PLANT PHYSICALS” section.  Rows beginning with
  ``Ore Milled (Tonnes)``, ``Mill Feed Grade (g/t)`` and
  ``Gold Recovered (kg)`` provide the monthly actuals.  Only the first
  numeric value on each row is used.  These values populate the ACTUAL
  section of the Monthly Stope Performance summary.

* **August Daily Report** –  The August daily report contains a sheet named
  ``Tramming``.  Within this sheet there is an “MTD/Budget/Actual/Variance”
  table where the column labelled “Actual” (Excel column D) holds the
  month‑to‑date figures for tonnage, grade and gold.  These numbers are
  used as the August actuals.  If the sheet structure differs, the script
  will ignore the daily report.

Once all available data have been extracted, the script updates the
Monthly Stope Performance workbook.  The summary sheet is searched for
the first occurrence of an ``ACTUAL`` row with ``Tonnes (t)``, and the two
rows below it are assumed to be ``Grade (g/t)`` and ``Gold (kg)``.  These
rows are overwritten for the months where actuals are supplied.  Similarly,
the first ``SHORT-TERM FORECAST`` row is located and the two rows below it
are updated with the forecast tonnages, grades and gold from the
3‑month rolling files.

The updated workbook is then returned to the user for download.
"""

import re
from datetime import datetime
from io import BytesIO
from typing import Dict, Iterable, Optional, Tuple

import openpyxl
import streamlit as st


def _extract_month_from_filename(filename: str) -> Optional[int]:
    """Attempt to infer the month index (1–12) from a 3‑month rolling filename.

    The naming convention assumed is "3 Months Rolling_<Month> <Year>.xlsx"
    (case insensitive).  Returns 1 for January through 12 for December,
    or ``None`` if parsing fails.
    """
    m = re.search(r"3\s*Months\s*Rolling[_\s-]*([A-Za-z]+)", filename)
    if not m:
        return None
    month_name = m.group(1).strip().lower()
    try:
        return datetime.strptime(month_name, "%B").month
    except ValueError:
        # try abbreviated month
        try:
            return datetime.strptime(month_name, "%b").month
        except ValueError:
            return None


def parse_three_month_rolling(file_bytes: BytesIO, filename: str) -> Optional[Tuple[float, float, float, int]]:
    """Parse a single 3‑month rolling workbook for forecast data.

    Returns a tuple ``(tonnes, grade, gold, month_index)`` or ``None`` if
    the necessary values cannot be extracted.
    The ``month_index`` corresponds to the month identified from the
    filename (January=1, etc.).
    """
    month_index = _extract_month_from_filename(filename)
    if month_index is None:
        return None
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    # Try to locate the sheet with milling plan; default to first sheet.
    sheet = wb.active
    if "MINE PLAN" in wb.sheetnames:
        sheet = wb["MINE PLAN"]
    # Search for the key rows within the sheet.
    tonne_value = grade_value = gold_value = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row or row[1] is None:
            continue
        label = str(row[1]).strip().lower()
        # Tonnes row immediately following a "MILLING" header.
        if label == "tonnes (t)":
            # The first numeric value in the row beyond column 2 is assumed to be
            # the tonnage for the file month.
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    tonne_value = float(val)
                    break
        elif label.startswith("grade"):
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    grade_value = float(val)
                    break
        elif label.startswith("3 month rolling"):
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    gold_value = float(val)
                    break
        if tonne_value is not None and grade_value is not None and gold_value is not None:
            break
    # Fallback: some files label gold as "gold (kg)"
    if gold_value is None:
        for row in sheet.iter_rows(values_only=True):
            if row and isinstance(row[1], str) and "gold" in row[1].lower():
                for val in row[2:]:
                    if isinstance(val, (int, float)):
                        gold_value = float(val)
                        break
                if gold_value is not None:
                    break
    if tonne_value is None or grade_value is None or gold_value is None:
        return None
    return tonne_value, grade_value, gold_value, month_index


def parse_actual_physical(file_bytes: BytesIO) -> Optional[Tuple[float, float, float, int]]:
    """Extract actuals from a single monthly physicals workbook.

    Returns ``(tonnes, grade, gold, month_index)`` or ``None`` if the
    sheet cannot be parsed.  The month index is taken from the date in
    cell B2 of the Dashboard (which stores the reporting period).  If the
    month cannot be determined, ``None`` is returned.
    """
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    if "Dashboard" not in wb.sheetnames:
        return None
    sheet = wb["Dashboard"]
    # Determine month from the reporting period (e.g., cell B2 contains the date)
    month_index = None
    for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
        for val in row:
            if isinstance(val, datetime):
                month_index = val.month
                break
        if month_index:
            break
    if not month_index:
        return None
    # Locate Plant Physicals rows
    tonne_value = grade_value = gold_value = None
    found_plant = False
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not found_plant and row[0] and "plant physicals" in str(row[0]).lower():
            found_plant = True
            continue
        if found_plant:
            label = row[0]
            if isinstance(label, str):
                low = label.lower().strip()
                if low.startswith("ore milled"):
                    for val in row[1:]:
                        if isinstance(val, (int, float)):
                            tonne_value = float(val)
                            break
                elif low.startswith("mill feed grade"):
                    for val in row[1:]:
                        if isinstance(val, (int, float)):
                            grade_value = float(val)
                            break
                elif low.startswith("gold recovered"):
                    for val in row[1:]:
                        if isinstance(val, (int, float)):
                            gold_value = float(val)
                            break
        if tonne_value is not None and grade_value is not None and gold_value is not None:
            break
    if tonne_value is None or grade_value is None or gold_value is None:
        return None
    return tonne_value, grade_value, gold_value, month_index


def parse_august_daily_report(file_bytes: BytesIO) -> Optional[Tuple[float, float, float]]:
    """Parse the August daily report and return the Actual MTD values.

    Looks for the ``Tramming`` sheet and reads the row labelled
    ``Trammed (t)``, ``Grade (g/t)``, and ``Gold (kg)`` from the column
    headed “Actual” (Excel column D).  Returns a tuple
    ``(tonnes, grade, gold)`` or ``None`` if parsing fails.
    """
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    if "Tramming" not in wb.sheetnames:
        return None
    sheet = wb["Tramming"]
    tonnes = grade = gold = None
    # Search for the summary table header row containing "MTD"
    header_row_index = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Row contains 'MTD' and 'Actual'
        if row and any(isinstance(val, str) and 'mtd' in val.lower() for val in row):
            header_row_index = i
            break
    if header_row_index is None:
        return None
    # Column index for Actual is typically 4 (D) – locate it relative to header
    header_row = [val if isinstance(val, str) else "" for val in sheet[header_row_index]]
    actual_col = None
    for idx, val in enumerate(header_row):
        if val.lower().strip() == 'actual':
            actual_col = idx + 1
            break
    if actual_col is None:
        # Fallback to column 5 (E)
        actual_col = 4
    # Now iterate the next few rows to find our metrics
    for row in sheet.iter_rows(min_row=header_row_index + 1, max_row=header_row_index + 10, values_only=False):
        label_cell = row[1].value if len(row) > 1 else None
        if isinstance(label_cell, str):
            low = label_cell.lower()
            if 'trammed' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    tonnes = float(val)
            elif 'grade' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    grade = float(val)
            elif 'gold' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    gold = float(val)
    if tonnes is None or grade is None or gold is None:
        return None
    return tonnes, grade, gold


def update_msp_workbook(msp_bytes: BytesIO,
                        forecasts: Dict[int, Tuple[float, float, float]],
                        actuals: Dict[int, Tuple[float, float, float]]) -> BytesIO:
    """Update the Monthly Stope Performance workbook with new forecasts and actuals.

    Parameters
    ----------
    msp_bytes : BytesIO
        The original workbook data as uploaded by the user.
    forecasts : dict
        Mapping of month index (1–12) to tuples of (tonnes, grade, gold)
        representing the short‑term forecast for that month.
    actuals : dict
        Mapping of month index (1–12) to tuples of (tonnes, grade, gold)
        representing the actual plant physicals for that month.

    Returns
    -------
    BytesIO
        The updated workbook encoded in memory.
    """
    wb = openpyxl.load_workbook(msp_bytes)
    if "SUMMARY" not in wb.sheetnames:
        raise ValueError("SUMMARY sheet not found in MSP workbook")
    ws = wb["SUMMARY"]
    # Locate the first forecast and actual rows
    forecast_start = actual_start = None
    for row_idx in range(1, ws.max_row + 1):
        row_label = ws.cell(row=row_idx, column=1).value
        row_type = ws.cell(row=row_idx, column=2).value
        if forecast_start is None and row_label == 'SHORT-TERM FORECAST' and row_type == 'Tonnes (t)':
            forecast_start = row_idx
        if actual_start is None and row_label == 'ACTUAL' and row_type == 'Tonnes (t)':
            actual_start = row_idx
        if forecast_start and actual_start:
            break
    if forecast_start is None or actual_start is None:
        raise ValueError("Could not locate forecast or actual rows in the SUMMARY sheet")
    # Rows for tonnes, grade, gold under forecast and actual
    forecast_rows = (forecast_start, forecast_start + 1, forecast_start + 2)
    actual_rows = (actual_start, actual_start + 1, actual_start + 2)
    # Starting column index for January (C = 3)
    jan_col = 3
    # Overwrite forecasts
    for month_idx, values in forecasts.items():
        col = jan_col + month_idx - 1
        tonnes, grade, gold = values
        ws.cell(row=forecast_rows[0], column=col).value = tonnes
        ws.cell(row=forecast_rows[1], column=col).value = grade
        ws.cell(row=forecast_rows[2], column=col).value = gold
    # Overwrite actuals
    for month_idx, values in actuals.items():
        col = jan_col + month_idx - 1
        tonnes, grade, gold = values
        ws.cell(row=actual_rows[0], column=col).value = tonnes
        ws.cell(row=actual_rows[1], column=col).value = grade
        ws.cell(row=actual_rows[2], column=col).value = gold
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def main() -> None:
    st.title("Monthly Stope Performance Updater")
    st.write(
        "Upload your existing Monthly Stope Performance workbook and any supporting\n"
        "files (3‑month rolling forecasts, monthly actual physicals reports, and\n"
        "an August daily report).  The application will extract the relevant metrics\n"
        "and update the SHORT‑TERM FORECAST and ACTUAL sections of the summary sheet."
    )
    msp_file = st.file_uploader(
        "Monthly Stope Performance file (required)", type=["xlsx"], key="msp"
    )
    three_month_files = st.file_uploader(
        "3‑Month Rolling files (optional, multiple)", type=["xlsx"], accept_multiple_files=True, key="three"
    )
    actual_files = st.file_uploader(
        "Actual Physicals files (optional, multiple)", type=["xlsx"], accept_multiple_files=True, key="actual"
    )
    daily_file = st.file_uploader(
        "August Daily Report (optional)", type=["xlsx"], key="daily"
    )
    if st.button("Update Workbook"):
        if msp_file is None:
            st.error("Please upload the Monthly Stope Performance workbook.")
            return
        forecasts: Dict[int, Tuple[float, float, float]] = {}
        # Parse 3‑month rolling files
        for uploaded in three_month_files:
            parsed = parse_three_month_rolling(BytesIO(uploaded.getvalue()), uploaded.name)
            if parsed:
                tonnes, grade, gold, month_idx = parsed
                forecasts[month_idx] = (tonnes, grade, gold)
        # Parse actuals
        actuals: Dict[int, Tuple[float, float, float]] = {}
        for uploaded in actual_files:
            parsed = parse_actual_physical(BytesIO(uploaded.getvalue()))
            if parsed:
                tonnes, grade, gold, month_idx = parsed
                actuals[month_idx] = (tonnes, grade, gold)
        # Parse August daily report
        if daily_file is not None:
            parsed = parse_august_daily_report(BytesIO(daily_file.getvalue()))
            if parsed:
                tonnes, grade, gold = parsed
                # August = 8
                actuals[8] = (tonnes, grade, gold)
        try:
            updated_bytes = update_msp_workbook(
                BytesIO(msp_file.getvalue()), forecasts, actuals
            )
        except Exception as exc:
            st.error(f"Failed to update workbook: {exc}")
            return
        st.success("Workbook updated successfully!")
        st.download_button(
            label="Download Updated Workbook",
            data=updated_bytes,
            file_name="Monthly Stope Performance Updated 5_08_2025.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()