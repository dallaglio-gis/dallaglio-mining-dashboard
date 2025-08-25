"""
Streamlit application to automate updating of the Monthly Stope Performance
workbook. Users can upload the primary workbook alongside a collection of
supporting spreadsheets (3‑month rolling forecasts, monthly actual physical
reports and an August daily report).  The tool parses the required metrics from
each source and overwrites the SHORT‑TERM FORECAST and ACTUAL sections in
the summary sheet of the Monthly Stope Performance file.

Parsing logic
-------------

* **3‑Month Rolling files** –  These workbooks typically contain a sheet
  named ``MINE PLAN`` where milling projections are tabulated.  The rows
  labelled “MILLING” and “Tonnes (t)” contain projected tonnages for the
  three months covered by the file.  Immediately below resides a row
  labelled “Grade - (g/t)”.  A separate row labelled “3 Month Rolling ”
  (note the trailing space) holds projected gold (kg).  The first numeric
  value in each of these rows corresponds to the month in the filename
  (e.g. ``3 Months Rolling_June 2025.xlsx`` → June).  Should this search
  fail, the parsing function returns ``None`` for that metric.

* **Actual Physicals files** –  For each monthly file the ``Dashboard``
  sheet contains a “PLANT PHYSICALS” section.  Rows beginning with
  ``Ore Milled (Tonnes)``, ``Mill Feed Grade (g/t)`` and
  ``Gold Recovered (kg)`` provide the monthly actuals.  Only the first
  numeric value on each row is used.  These values populate the ACTUAL
  section of the Monthly Stope Performance summary.

* **August Daily Report** –  The August daily report contains a sheet named
  ``Tramming``.  Within this sheet there is an “MTD/Budget/Actual/Variance”
  table where the column labelled “Actual” (Excel column D) holds the
  month‑to‑date figures for tonnage, grade and gold.  These numbers are
  used as the August actuals.  If the sheet structure differs, the script
  will ignore the daily report.

Once all available data have been extracted, the script updates the
Monthly Stope Performance workbook.  The summary sheet is searched for
the first occurrence of an ``ACTUAL`` row with ``Tonnes (t)``, and the two
rows below it are assumed to be ``Grade (g/t)`` and ``Gold (kg)``.  These
rows are overwritten for the months where actuals are supplied.  Similarly,
the first ``SHORT-TERM FORECAST`` row is located and the two rows below it
are updated with the forecast tonnages, grades and gold from the
3‑month rolling files.

The updated workbook is then returned to the user for download.
"""

import re
from datetime import datetime
from io import BytesIO
from typing import Dict, Iterable, Optional, Tuple, List, Set

import openpyxl
import streamlit as st

###############################################################################
# Utility functions for ID normalisation and month calculations
###############################################################################

def _normalize_id(id_str: str) -> str:
    """Normalise stope IDs by replacing underscores with spaces, collapsing
    multiple spaces, stripping whitespace and converting to uppercase.

    Many of the source files use slightly different naming conventions
    (e.g. "1L_W28_STOPE" versus "1L W28 STOPE").  This helper reduces those
    inconsistencies so IDs can be matched reliably across different sheets.
    """
    if not id_str:
        return ""
    cleaned = id_str.replace("_", " ").replace("\n", " ")
    cleaned = " ".join(cleaned.split())
    return cleaned.strip().upper()

def _month_add(base_month: int, offset: int) -> int:
    """Return the month index after adding an offset.  Keeps result in 1–12.

    Example: ``_month_add(11, 2) -> 1`` (January of next year).
    """
    return ((base_month - 1 + offset) % 12) + 1


def _extract_month_from_filename(filename: str) -> Optional[int]:
    """Attempt to infer the month index (1–12) from a 3‑month rolling filename.

    The naming convention assumed is "3 Months Rolling_<Month> <Year>.xlsx"
    (case insensitive).  Returns 1 for January through 12 for December,
    or ``None`` if parsing fails.
    """
    m = re.search(r"3\s*Months\s*Rolling[_\s-]*([A-Za-z]+)", filename)
    if not m:
        return None
    month_name = m.group(1).strip().lower()
    try:
        return datetime.strptime(month_name, "%B").month
    except ValueError:
        # try abbreviated month
        try:
            return datetime.strptime(month_name, "%b").month
        except ValueError:
            return None


def parse_three_month_rolling(file_bytes: BytesIO, filename: str) -> Optional[Dict[int, Tuple[float, float, float]]]:
    """Parse a 3‑month rolling forecast workbook for forecast data.

    Returns a mapping of month indices (1–12) to tuples ``(tonnes, grade, gold)``.
    The first month is inferred from the filename and subsequent two months
    follow sequentially.  If the workbook structure cannot be parsed,
    ``None`` is returned.

    In most files the rows ``Tonnes (t)``, ``Grade - (g/t)`` and ``3 Month Rolling ``
    (or occasionally ``Gold (kg)``) hold the projections for the three months
    covered by the file.  The parser will extract up to three numeric values
    from each of these rows and assign them to the appropriate month.
    """
    start_month = _extract_month_from_filename(filename)
    if start_month is None:
        return None
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    sheet = wb.active
    if "MINE PLAN" in wb.sheetnames:
        sheet = wb["MINE PLAN"]
    # Collect numeric values for each metric
    tonnes: List[float] = []
    grades: List[float] = []
    golds: List[float] = []
    for row in sheet.iter_rows(values_only=True):
        if not row or len(row) < 2 or row[1] is None:
            continue
        label = str(row[1]).strip().lower()
        if label == "tonnes (t)":
            vals: List[float] = []
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    vals.append(float(val))
                if len(vals) >= 3:
                    break
            if vals:
                tonnes = vals
        elif label.startswith("grade"):
            vals: List[float] = []
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    vals.append(float(val))
                if len(vals) >= 3:
                    break
            if vals:
                grades = vals
        elif label.startswith("3 month rolling") or "gold" in label:
            vals: List[float] = []
            for val in row[2:]:
                if isinstance(val, (int, float)):
                    vals.append(float(val))
                if len(vals) >= 3:
                    break
            if vals:
                golds = vals
        if tonnes and grades and golds:
            break
    # Fallback search for gold values if none captured
    if not golds:
        for row in sheet.iter_rows(values_only=True):
            if row and len(row) > 1 and isinstance(row[1], str) and "gold" in row[1].lower():
                vals: List[float] = []
                for val in row[2:]:
                    if isinstance(val, (int, float)):
                        vals.append(float(val))
                    if len(vals) >= 3:
                        break
                if vals:
                    golds = vals
                    break
    if not tonnes or not grades or not golds:
        return None
    # Pad lists to length 3
    while len(tonnes) < 3:
        tonnes.append(0.0)
    while len(grades) < 3:
        grades.append(0.0)
    while len(golds) < 3:
        golds.append(0.0)
    forecasts: Dict[int, Tuple[float, float, float]] = {}
    for offset in range(3):
        month = _month_add(start_month, offset)
        forecasts[month] = (tonnes[offset], grades[offset], golds[offset])
    return forecasts


def parse_actual_physical(file_bytes: BytesIO) -> Optional[Tuple[float, float, float, int]]:
    """Extract actuals from a single monthly physicals workbook.

    Returns ``(tonnes, grade, gold, month_index)`` or ``None`` if the
    sheet cannot be parsed.  The month index is taken from the date in
    cell B2 of the Dashboard (which stores the reporting period).  If the
    month cannot be determined, ``None`` is returned.
    """
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    if "Dashboard" not in wb.sheetnames:
        return None
    sheet = wb["Dashboard"]
    # Determine month from the reporting period (e.g., cell B2 contains the date)
    month_index = None
    for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
        for val in row:
            if isinstance(val, datetime):
                month_index = val.month
                break
        if month_index:
            break
    if not month_index:
        return None
    # Locate Plant Physicals rows
    tonne_value = grade_value = gold_value = None
    found_plant = False
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not found_plant and row[0] and "plant physicals" in str(row[0]).lower():
            found_plant = True
            continue
        if found_plant:
            label = row[0]
            if isinstance(label, str):
                low = label.lower().strip()
                if low.startswith("ore milled"):
                    for val in row[1:]:
                        if isinstance(val, (int, float)):
                            tonne_value = float(val)
                            break
                elif low.startswith("mill feed grade"):
                    for val in row[1:]:
                        if isinstance(val, (int, float)):
                            grade_value = float(val)
                            break
                elif low.startswith("gold recovered"):
                    for val in row[1:]:
                        if isinstance(val, (int, float)):
                            gold_value = float(val)
                            break
        if tonne_value is not None and grade_value is not None and gold_value is not None:
            break
    if tonne_value is None or grade_value is None or gold_value is None:
        return None
    return tonne_value, grade_value, gold_value, month_index


def parse_august_daily_report(file_bytes: BytesIO) -> Optional[Tuple[float, float, float]]:
    """Parse the August daily report and return the Actual MTD values.

    Looks for the ``Tramming`` sheet and reads the row labelled
    ``Trammed (t)``, ``Grade (g/t)``, and ``Gold (kg)`` from the column
    headed “Actual” (Excel column D).  Returns a tuple
    ``(tonnes, grade, gold)`` or ``None`` if parsing fails.
    """
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    if "Tramming" not in wb.sheetnames:
        return None
    sheet = wb["Tramming"]
    tonnes = grade = gold = None
    # Search for the summary table header row containing "MTD"
    header_row_index = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Row contains 'MTD' and 'Actual'
        if row and any(isinstance(val, str) and 'mtd' in val.lower() for val in row):
            header_row_index = i
            break
    if header_row_index is None:
        return None
    # Column index for Actual is typically 4 (D) – locate it relative to header
    header_row = [val if isinstance(val, str) else "" for val in sheet[header_row_index]]
    actual_col = None
    for idx, val in enumerate(header_row):
        if val.lower().strip() == 'actual':
            actual_col = idx + 1
            break
    if actual_col is None:
        # Fallback to column 4 (D)
        actual_col = 4
    # Now iterate the next few rows to find our metrics
    for row in sheet.iter_rows(min_row=header_row_index + 1, max_row=header_row_index + 10, values_only=False):
        label_cell = row[1].value if len(row) > 1 else None
        if isinstance(label_cell, str):
            low = label_cell.lower()
            if 'trammed' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    tonnes = float(val)
            elif 'grade' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    grade = float(val)
            elif 'gold' in low:
                val = row[actual_col - 1].value
                if isinstance(val, (int, float)):
                    gold = float(val)
    if tonnes is None or grade is None or gold is None:
        return None
    return tonnes, grade, gold

###############################################################################
# New parsing functions for Planned Not Mined (PNM) and Mined Not Planned (MNP)
###############################################################################

def parse_underground_breaking_plan(file_bytes: BytesIO) -> Set[str]:
    """Extract stope IDs from an Underground Breaking Plan workbook.

    Mining & Milling Plan files contain an ``Underground Breaking Plan`` sheet
    listing all stopes scheduled for blasting.  Each stope appears on a row
    where the second column (index 1) contains the label ``Stoping Tons``.
    The first column (index 0) of such rows holds the stope identifier.  This
    function normalises those identifiers and returns a set of unique IDs.

    Parameters
    ----------
    file_bytes : BytesIO
        The raw bytes of a Mining & Milling Plan workbook.

    Returns
    -------
    set of str
        A set of normalised stope IDs scheduled for breaking.
    """
    ids: Set[str] = set()
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return ids
    # Try to locate the sheet by name (case insensitive)
    target_name = None
    for name in wb.sheetnames:
        if 'underground breaking plan' in name.lower():
            target_name = name
            break
    if not target_name:
        return ids
    ws = wb[target_name]
    for row in ws.iter_rows(values_only=True):
        # Each stope row has 'Stoping Tons' in the second column
        if row and len(row) > 1:
            second = row[1]
            if isinstance(second, str) and second.strip().lower() == 'stoping tons':
                stope = row[0]
                if isinstance(stope, str):
                    ids.add(_normalize_id(stope))
    return ids


def parse_tramming_detail(file_bytes: BytesIO) -> Optional[Tuple[int, Dict[str, Tuple[float, float, float, float, float, float]]]]:
    """Parse a detailed tramming daily report for stope-level budgets and actuals.

    The tramming sheets of the daily reports contain repeated blocks for each
    stope.  Each block begins with a row where one of the cells contains
    ``STOPE`` (e.g. ``1L_W28_STOPE``).  Subsequent rows within the block
    describe ``Tonnes``, ``Grade`` and ``Gold`` metrics, each with two
    sub‑rows: ``Budget`` and ``Actual``.  The first numeric value to the right
    of ``Budget (t)``, ``Budget (g/t)`` or ``Budget (kg)`` is taken as the
    month‑to‑date budget for that metric.  Likewise, the first numeric value
    to the right of ``Actual (t)``, ``Actual (g/t)`` or ``Actual (kg)`` is the
    month‑to‑date actual.  Rows labelled ``Reason if Actual is RED`` are
    ignored.

    The month of the report is determined by searching for a cell containing
    ``Month:`` with an adjacent datetime value.  The resulting month index
    (1–12) is returned along with a mapping from stope ID to a tuple of six
    floats: ``(budget_tonnes, actual_tonnes, budget_grade, actual_grade,
    budget_gold, actual_gold)``.

    Parameters
    ----------
    file_bytes : BytesIO
        Raw bytes of a daily report workbook containing a ``Tramming`` sheet.

    Returns
    -------
    tuple
        ``(month_index, data)`` where ``data`` maps normalised stope IDs
        to their budget and actual metrics.  Returns ``None`` if the sheet
        cannot be parsed.
    """
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        return None
    # Identify the tramming sheet (the first sheet containing 'tramming' in its name)
    sheet_name = None
    for name in wb.sheetnames:
        if 'tramming' in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        return None
    ws = wb[sheet_name]
    # Determine the reporting month
    month_idx: Optional[int] = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        for i, val in enumerate(row):
            if isinstance(val, str) and 'month' in val.lower():
                # Expect adjacent cell contains a date
                if i + 1 < len(row) and isinstance(row[i + 1], datetime):
                    month_idx = row[i + 1].month
                    break
        if month_idx:
            break
    if not month_idx:
        return None
    data: Dict[str, Tuple[float, float, float, float, float, float]] = {}
    rows = list(ws.iter_rows(values_only=True))
    n = len(rows)
    i = 0
    while i < n:
        row = rows[i]
        # Look for a stope identifier
        stope_id = None
        for cell in row:
            if isinstance(cell, str) and 'stope' in cell.lower():
                stope_id = _normalize_id(cell)
                break
        if stope_id:
            # Initialise placeholders
            budget_tonnes = actual_tonnes = 0.0
            budget_grade = actual_grade = 0.0
            budget_gold = actual_gold = 0.0
            # Scan subsequent rows until next stope or end
            j = i + 1
            while j < n:
                next_row = rows[j]
                # If next row contains another stope, break
                if any(isinstance(c, str) and 'stope' in c.lower() for c in next_row if c):
                    break
                # Extract budgets and actuals
                if next_row:
                    for k, val in enumerate(next_row):
                        if isinstance(val, str):
                            label = val.strip().lower()
                            # Budget tonnage
                            if label == 'budget (t)':
                                # find first numeric cell to the right
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        budget_tonnes = float(x)
                                        break
                            elif label == 'actual (t)':
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        actual_tonnes = float(x)
                                        break
                            elif label in ('budget (g/t)', 'budget (gpt)', 'budget (gpt)', 'budget (g/t)'):
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        budget_grade = float(x)
                                        break
                            elif label in ('actual (g/t)', 'actual (gpt)', 'actual (gpt)', 'actual (g/t)'):
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        actual_grade = float(x)
                                        break
                            elif label == 'budget (kg)':
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        budget_gold = float(x)
                                        break
                            elif label == 'actual (kg)':
                                for x in next_row[k + 1:]:
                                    if isinstance(x, (int, float)):
                                        actual_gold = float(x)
                                        break
                j += 1
            data[stope_id] = (
                budget_tonnes,
                actual_tonnes,
                budget_grade,
                actual_grade,
                budget_gold,
                actual_gold,
            )
            # Continue scanning from the next row after this block
            i = j
            continue
        i += 1
    return month_idx, data


def update_pnm_mnp_sheet(wb: openpyxl.Workbook,
                         pnm_data: Dict[str, Tuple[float, float, float]],
                         mnp_data: Dict[str, Tuple[float, float, float]],
                         month_idx: int) -> None:
    """Update the ``Stopes PNM & MNP`` sheet with new PNM and MNP values.

    This function modifies the workbook in place.  It determines the column
    positions corresponding to the given month index, inserts new rows as
    necessary for previously unseen stopes and recalculates total values for
    the month across all stopes.

    Parameters
    ----------
    wb : openpyxl.Workbook
        The workbook object returned by ``update_msp_workbook``.
    pnm_data : dict
        Mapping of stope ID to a tuple ``(diff_tonnes, diff_grade, diff_gold)``
        representing Planned Not Mined metrics for the month.
    mnp_data : dict
        Mapping of stope ID to a tuple ``(actual_tonnes, actual_grade,
        actual_gold)`` representing Mined Not Planned metrics for the month.
    month_idx : int
        The month index (1–12) corresponding to the daily report.
    """
    if 'Stopes PNM & MNP' not in wb.sheetnames:
        return
    ws = wb['Stopes PNM & MNP']
    # Helper to build month column mapping for both sections.  Row containing
    # month dates is the one under each section header (e.g. row with datetime cells).
    def build_month_map(start_row: int) -> Dict[int, int]:
        """Return mapping of month index to starting column for a section."""
        month_map: Dict[int, int] = {}
        # row at start_row + 1 holds date headers (Tonnes/Grade/Gold names below)
        header_row = ws[start_row + 1]
        for col_idx, cell in enumerate(header_row, start=1):
            if isinstance(cell.value, datetime):
                month_map[cell.value.month] = col_idx
        return month_map
    # Locate PNM and MNP section starts
    pnm_start = mnp_start = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[1] and isinstance(row[1], str):
            val = row[1].strip().lower()
            if val == 'planned not mined' and pnm_start is None:
                pnm_start = idx
            elif val == 'mined not planned' and mnp_start is None:
                mnp_start = idx
        if pnm_start and mnp_start:
            break
    if pnm_start is None or mnp_start is None:
        return
    # Build month column maps
    pnm_month_map = build_month_map(pnm_start + 1)
    mnp_month_map = build_month_map(mnp_start + 1)
    # Determine start columns for this month (skip if not present)
    pnm_col_start = pnm_month_map.get(month_idx)
    mnp_col_start = mnp_month_map.get(month_idx)
    # If month not present in sheet, do nothing
    if not pnm_col_start and not mnp_col_start:
        return
    # Identify the first stope row and total row indices for each section
    def find_stope_rows(section_start: int) -> Tuple[int, int]:
        """Find start and end row indices (inclusive) of stope data for a section."""
        start = section_start + 3  # two header rows and one subheader row
        # iterate until we hit a row where column 2 is 'Total' or empty
        end = start - 1
        for r in range(start, ws.max_row + 1):
            val = ws.cell(row=r, column=2).value
            if val is None:
                break
            if isinstance(val, str) and val.strip().lower() == 'total':
                end = r - 1
                break
        return start, end
    pnm_stope_start, pnm_stope_end = find_stope_rows(pnm_start)
    mnp_stope_start, mnp_stope_end = find_stope_rows(mnp_start)
    # Helper to ensure a stope row exists; insert if not
    def get_or_create_row(section_start: int, stope_start: int, stope_end: int, stope_id: str) -> int:
        # Search existing rows for the stope ID
        for r in range(stope_start, stope_end + 1):
            val = ws.cell(row=r, column=2).value
            if isinstance(val, str) and _normalize_id(val) == stope_id:
                return r
        # Insert a new row before the total row (stope_end + 1 is the first total row)
        insert_row = stope_end + 1
        ws.insert_rows(insert_row)
        # Write the stope ID into column 2
        ws.cell(row=insert_row, column=2, value=stope_id)
        # Return the new row index
        return insert_row
    # Update PNM section
    if pnm_col_start:
        for stope_id, vals in pnm_data.items():
            diff_tonnes, diff_grade, diff_gold = vals
            row_idx = get_or_create_row(pnm_start, pnm_stope_start, pnm_stope_end, stope_id)
            ws.cell(row=row_idx, column=pnm_col_start, value=diff_tonnes)
            ws.cell(row=row_idx, column=pnm_col_start + 1, value=diff_grade)
            ws.cell(row=row_idx, column=pnm_col_start + 2, value=diff_gold)
            # Update PNM stope_end if a new row was added
            if row_idx > pnm_stope_end:
                pnm_stope_end = row_idx
        # Recalculate totals for PNM
        # Locate the total row (the first row where column 2 is 'Total' after start)
        total_row = pnm_stope_end + 1
        # Compute per-column totals for the month
        # Tonnes
        tot_tonnes = 0.0
        weighted_grade_sum = 0.0
        tot_gold = 0.0
        for r in range(pnm_stope_start, pnm_stope_end + 1):
            v_t = ws.cell(row=r, column=pnm_col_start).value or 0
            v_g = ws.cell(row=r, column=pnm_col_start + 1).value or 0
            v_au = ws.cell(row=r, column=pnm_col_start + 2).value or 0
            # convert to float where possible
            try:
                v_t = float(v_t)
            except Exception:
                v_t = 0.0
            try:
                v_g = float(v_g)
            except Exception:
                v_g = 0.0
            try:
                v_au = float(v_au)
            except Exception:
                v_au = 0.0
            tot_tonnes += v_t
            weighted_grade_sum += v_t * v_g
            tot_gold += v_au
        # Write totals
        ws.cell(row=total_row, column=pnm_col_start, value=tot_tonnes)
        # Weighted average grade
        avg_grade = weighted_grade_sum / tot_tonnes if tot_tonnes else 0
        ws.cell(row=total_row, column=pnm_col_start + 1, value=avg_grade)
        ws.cell(row=total_row, column=pnm_col_start + 2, value=tot_gold)
    # Update MNP section
    if mnp_col_start:
        for stope_id, vals in mnp_data.items():
            act_tonnes, act_grade, act_gold = vals
            row_idx = get_or_create_row(mnp_start, mnp_stope_start, mnp_stope_end, stope_id)
            ws.cell(row=row_idx, column=mnp_col_start, value=act_tonnes)
            ws.cell(row=row_idx, column=mnp_col_start + 1, value=act_grade)
            ws.cell(row=row_idx, column=mnp_col_start + 2, value=act_gold)
            # Update MNP stope_end if a new row was added
            if row_idx > mnp_stope_end:
                mnp_stope_end = row_idx
        # Recalculate totals for MNP
        total_row = mnp_stope_end + 1
        tot_tonnes = 0.0
        weighted_grade_sum = 0.0
        tot_gold = 0.0
        for r in range(mnp_stope_start, mnp_stope_end + 1):
            v_t = ws.cell(row=r, column=mnp_col_start).value or 0
            v_g = ws.cell(row=r, column=mnp_col_start + 1).value or 0
            v_au = ws.cell(row=r, column=mnp_col_start + 2).value or 0
            try:
                v_t = float(v_t)
            except Exception:
                v_t = 0.0
            try:
                v_g = float(v_g)
            except Exception:
                v_g = 0.0
            try:
                v_au = float(v_au)
            except Exception:
                v_au = 0.0
            tot_tonnes += v_t
            weighted_grade_sum += v_t * v_g
            tot_gold += v_au
        ws.cell(row=total_row, column=mnp_col_start, value=tot_tonnes)
        avg_grade = weighted_grade_sum / tot_tonnes if tot_tonnes else 0
        ws.cell(row=total_row, column=mnp_col_start + 1, value=avg_grade)
        ws.cell(row=total_row, column=mnp_col_start + 2, value=tot_gold)


def update_msp_workbook(msp_bytes: BytesIO,
                        forecasts: Dict[int, Tuple[float, float, float]],
                        actuals: Dict[int, Tuple[float, float, float]]) -> BytesIO:
    """Update the Monthly Stope Performance workbook with new forecasts and actuals.

    Parameters
    ----------
    msp_bytes : BytesIO
        The original workbook data as uploaded by the user.
    forecasts : dict
        Mapping of month index (1–12) to tuples of (tonnes, grade, gold)
        representing the short‑term forecast for that month.
    actuals : dict
        Mapping of month index (1–12) to tuples of (tonnes, grade, gold)
        representing the actual plant physicals for that month.

    Returns
    -------
    BytesIO
        The updated workbook encoded in memory.
    """
    wb = openpyxl.load_workbook(msp_bytes)
    if "SUMMARY" not in wb.sheetnames:
        raise ValueError("SUMMARY sheet not found in MSP workbook")
    ws = wb["SUMMARY"]
    # Locate the first forecast and actual rows
    forecast_start = actual_start = None
    for row_idx in range(1, ws.max_row + 1):
        row_label = ws.cell(row=row_idx, column=1).value
        row_type = ws.cell(row=row_idx, column=2).value
        if forecast_start is None and row_label == 'SHORT-TERM FORECAST' and row_type == 'Tonnes (t)':
            forecast_start = row_idx
        if actual_start is None and row_label == 'ACTUAL' and row_type == 'Tonnes (t)':
            actual_start = row_idx
        if forecast_start and actual_start:
            break
    if forecast_start is None or actual_start is None:
        raise ValueError("Could not locate forecast or actual rows in the SUMMARY sheet")
    # Rows for tonnes, grade, gold under forecast and actual
    forecast_rows = (forecast_start, forecast_start + 1, forecast_start + 2)
    actual_rows = (actual_start, actual_start + 1, actual_start + 2)
    # Starting column index for January (C = 3)
    jan_col = 3
    # Overwrite forecasts
    for month_idx, values in forecasts.items():
        col = jan_col + month_idx - 1
        tonnes, grade, gold = values
        ws.cell(row=forecast_rows[0], column=col).value = tonnes
        ws.cell(row=forecast_rows[1], column=col).value = grade
        ws.cell(row=forecast_rows[2], column=col).value = gold
    # Overwrite actuals
    for month_idx, values in actuals.items():
        col = jan_col + month_idx - 1
        tonnes, grade, gold = values
        ws.cell(row=actual_rows[0], column=col).value = tonnes
        ws.cell(row=actual_rows[1], column=col).value = grade
        ws.cell(row=actual_rows[2], column=col).value = gold
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def main() -> None:
    st.title("Monthly Stope Performance Updater")
    st.write(
        "Upload your existing Monthly Stope Performance workbook and any supporting\n"
        "files (3‑month rolling forecasts, monthly actual physicals reports, and\n"
        "an August daily report).  The application will extract the relevant metrics\n"
        "and update the SHORT‑TERM FORECAST and ACTUAL sections of the summary sheet."
    )
    msp_file = st.file_uploader(
        "Monthly Stope Performance file (required)", type=["xlsx"], key="msp"
    )
    three_month_files = st.file_uploader(
        "3‑Month Rolling files (optional, multiple)",
        type=["xlsx"], accept_multiple_files=True, key="three"
    )
    actual_files = st.file_uploader(
        "Actual Physicals files (optional, multiple)",
        type=["xlsx"], accept_multiple_files=True, key="actual"
    )
    # Mining & Milling plan files provide the underground breaking plan
    plan_files = st.file_uploader(
        "Mining & Milling Plan files (optional, multiple)",
        type=["xlsx"], accept_multiple_files=True, key="plan"
    )
    # Detailed tramming daily reports containing stope budgets and actuals
    tramming_files = st.file_uploader(
        "Tramming Daily Report files (optional, multiple)",
        type=["xlsx"], accept_multiple_files=True, key="tramming"
    )
    # Simple daily report for August used for summary actuals (retain original functionality)
    daily_file = st.file_uploader(
        "August Daily Report (optional)", type=["xlsx"], key="daily"
    )
    if st.button("Update Workbook"):
        if msp_file is None:
            st.error("Please upload the Monthly Stope Performance workbook.")
            return
        forecasts: Dict[int, Tuple[float, float, float]] = {}
        # Parse 3‑month rolling files.  Each file may contain forecasts for up to
        # three months.  We merge them into the forecasts dict, later entries
        # overwrite earlier ones for the same month.
        for uploaded in three_month_files:
            parsed = parse_three_month_rolling(BytesIO(uploaded.getvalue()), uploaded.name)
            if parsed:
                for month_idx, triple in parsed.items():
                    forecasts[month_idx] = triple
        # Parse actuals
        actuals: Dict[int, Tuple[float, float, float]] = {}
        for uploaded in actual_files:
            parsed = parse_actual_physical(BytesIO(uploaded.getvalue()))
            if parsed:
                tonnes, grade, gold, month_idx = parsed
                actuals[month_idx] = (tonnes, grade, gold)
        # Parse simple August daily report for summary actuals
        if daily_file is not None:
            parsed = parse_august_daily_report(BytesIO(daily_file.getvalue()))
            if parsed:
                tonnes, grade, gold = parsed
                # August = 8
                actuals[8] = (tonnes, grade, gold)
        # Parse mining & milling plan files to build a set of planned stope IDs
        planned_ids: Set[str] = set()
        for uploaded in plan_files:
            ids = parse_underground_breaking_plan(BytesIO(uploaded.getvalue()))
            planned_ids.update(ids)
        # Parse detailed tramming reports to compute PNM and MNP data per month
        pnm_by_month: Dict[int, Dict[str, Tuple[float, float, float]]] = {}
        mnp_by_month: Dict[int, Dict[str, Tuple[float, float, float]]] = {}
        for uploaded in tramming_files:
            res = parse_tramming_detail(BytesIO(uploaded.getvalue()))
            if res:
                month_idx, tramming_data = res
                pnm_data: Dict[str, Tuple[float, float, float]] = {}
                mnp_data: Dict[str, Tuple[float, float, float]] = {}
                for stope_id, (b_t, a_t, b_g, a_g, b_au, a_au) in tramming_data.items():
                    diff_t = (b_t or 0) - (a_t or 0)
                    diff_g = (b_g or 0) - (a_g or 0)
                    diff_au = (b_au or 0) - (a_au or 0)
                    # Only record if any difference exists
                    if diff_t != 0 or diff_g != 0 or diff_au != 0:
                        pnm_data[stope_id] = (diff_t, diff_g, diff_au)
                    # MNP: stope mined but not planned
                    if stope_id not in planned_ids:
                        mnp_data[stope_id] = (a_t or 0, a_g or 0, a_au or 0)
                # Merge with existing month data if multiple files per month
                if month_idx in pnm_by_month:
                    pnm_by_month[month_idx].update(pnm_data)
                else:
                    pnm_by_month[month_idx] = pnm_data
                if month_idx in mnp_by_month:
                    mnp_by_month[month_idx].update(mnp_data)
                else:
                    mnp_by_month[month_idx] = mnp_data
        # Update the summary forecast and actual sections
        try:
            updated_bytes = update_msp_workbook(
                BytesIO(msp_file.getvalue()), forecasts, actuals
            )
        except Exception as exc:
            st.error(f"Failed to update workbook: {exc}")
            return
        # Open the workbook for additional PNM/MNP updates
        try:
            wb = openpyxl.load_workbook(updated_bytes)
        except Exception as exc:
            st.error(f"Failed to open updated workbook for PNM/MNP updates: {exc}")
            return
        # Apply PNM and MNP updates per month
        for month_idx, pnm_data in pnm_by_month.items():
            mnp_data = mnp_by_month.get(month_idx, {})
            update_pnm_mnp_sheet(wb, pnm_data, mnp_data, month_idx)
        # Save final workbook
        final_bytes = BytesIO()
        wb.save(final_bytes)
        final_bytes.seek(0)
        st.success("Workbook updated successfully!")
        st.download_button(
            label="Download Updated Workbook",
            data=final_bytes,
            file_name="Monthly Stope Performance 05_08.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()