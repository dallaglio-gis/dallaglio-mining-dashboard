
"""
HOISTING Sheet Data Processor
Handles the unique Source/METRIC1/METRIC2/Value format
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Tuple, List, Dict
import sys
import os

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# NOTE: We avoid importing project specific utilities such as clean_and_validate_data,
# setup_logger and detect_date_columns because they may not be available in all
# execution contexts. Instead we implement minimal helpers inline and fall
# back to basic Python logging.
from openpyxl import load_workbook
import logging

def setup_logger(name: str, log_file: str, level: int = logging.INFO):
    """Set up a simple logger. If a file cannot be written the logger will
    log to the console instead."""
    logger = logging.getLogger(name)
    if logger.handlers:
        return logger
    logger.setLevel(level)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    try:
        fh = logging.FileHandler(log_file)
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    except Exception:
        ch = logging.StreamHandler()
        ch.setFormatter(formatter)
        logger.addHandler(ch)
    return logger

def clean_value(value):
    """Convert cell values to float where appropriate. Returns 0 for None or
    non-numeric entries."""
    if value is None:
        return 0.0
    try:
        # Convert boolean or string numbers to float
        return float(value)
    except Exception:
        return 0.0

class HoistingProcessor:
    """Processor for HOISTING sheet data extraction"""
    
    def __init__(self, logger=None):
        self.logger = logger or setup_logger('HoistingProcessor', 'logs/hoisting.log')
    
    def extract_hoisting_data(self, file_path: str) -> pd.DataFrame:
        """Extract hoisting daily actual and budget figures for tonnes, grade and gold.

        The Hoisting worksheet arranges daily budgets and actuals horizontally
        starting at the column immediately to the right of the "Shifts" label. The
        row containing this label also provides the day numbers. Subsequent rows
        contain daily budget and actual figures for tonnes (t), grade (g/t) and
        gold (kg). This method reads those values, constructs proper dates using
        the month specified on the sheet, and returns a flat table with one
        record per day. Missing budgets or actuals are replaced with 0.
        """
        self.logger.info("Loading HOISTING sheet...")
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb['Hoisting']
        except Exception as e:
            self.logger.error(f"Failed to load HOISTING sheet: {e}")
            return pd.DataFrame()

        # Determine the start date (year and month) for the sheet. Look for the
        # 'Month:' label and read the cell to its right which should contain
        # the first day of the month as a datetime. If not found, we'll leave
        # month_start_date as None and the Date column will be None.
        month_start_date = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower().startswith('month'):
                    # openpyxl's Cell.offset takes positional arguments (row, column)
                    candidate = cell.offset(0, 1).value
                    if isinstance(candidate, datetime):
                        month_start_date = candidate
                    break
            if month_start_date:
                break

        if month_start_date is None:
            self.logger.warning("Month start date not found in Hoisting sheet; dates will be None")

        # Locate the row and column of the 'Shifts' cell. This row also
        # contains the day numbers starting in the next column.
        shifts_row_idx = None
        shifts_col_idx = None
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == 'shifts':
                    shifts_row_idx = cell.row
                    shifts_col_idx = cell.column
                    break
            if shifts_row_idx is not None:
                break

        if shifts_row_idx is None:
            self.logger.error("'Shifts' label not found in Hoisting sheet")
            return pd.DataFrame()

        # Days start in the column immediately after 'Shifts'
        day_start_col = shifts_col_idx + 1
        day_nums: List[int] = []
        col = day_start_col
        # Iterate through the header row to collect day numbers until an empty cell
        while True:
            val = ws.cell(row=shifts_row_idx, column=col).value
            if val is None:
                break
            try:
                day_int = int(val)
                day_nums.append(day_int)
            except Exception:
                break
            col += 1

        if not day_nums:
            self.logger.error("No day numbers found on Hoisting sheet")
            return pd.DataFrame()

        # Define the labels we expect in the sheet for each data series. Map
        # internal keys to the literal text found in the sheet. These labels
        # reside in the column immediately to the left of the daily values.
        search_labels = {
            'budget_t': 'Daily Budget (t)',
            'actual_t': 'Daily Actual (t)',
            'budget_grade': 'Daily Budget (g/t)',
            'actual_grade': 'Daily Actual (g/t)',
            'budget_gold': 'Daily Budget (kg)',
            'actual_gold': 'Daily Actual (kg)'
        }

        # Locate the row index for each label. We'll scan the sheet and record
        # the row where the exact label appears. If a particular label is not
        # found, we'll leave its row index as None.
        label_rows: Dict[str, int] = {key: None for key in search_labels}
        for row in ws.iter_rows(min_row=1, max_row=30):
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    stripped = val.strip()
                    for key, label in search_labels.items():
                        if stripped == label:
                            label_rows[key] = cell.row
                    # If we've found all labels, we can stop searching
            if all(v is not None for v in label_rows.values()):
                break

        # Prepare containers for each series with default zeros
        data_dict: Dict[str, List[float]] = {}
        for key in search_labels:
            data_dict[key] = [0.0] * len(day_nums)

        # Extract the data across the day columns for each found row
        for key, row_idx in label_rows.items():
            if row_idx is None:
                # Keep default zeros if the row isn't present
                continue
            for idx, col_offset in enumerate(range(day_start_col, day_start_col + len(day_nums))):
                cell_val = ws.cell(row=row_idx, column=col_offset).value
                data_dict[key][idx] = clean_value(cell_val)

        # Construct proper dates if a start date is available
        date_list: List[datetime] = []
        if isinstance(month_start_date, datetime):
            year = month_start_date.year
            month = month_start_date.month
            for d in day_nums:
                try:
                    date_list.append(datetime(year, month, int(d)))
                except Exception:
                    date_list.append(None)
        else:
            date_list = [None] * len(day_nums)

        # Build a flat table with the required columns. Use the names
        # prescribed in the user instructions.
        records: List[Dict[str, object]] = []
        for idx, date_val in enumerate(date_list):
            rec = {
                'Date': date_val,
                'Hoisted_Actual_t': data_dict['actual_t'][idx],
                'Hoisting_Actual_gpt': data_dict['actual_grade'][idx],
                'Hoisted_Actual_kg': data_dict['actual_gold'][idx],
                'Hoisted_Budget_t': data_dict['budget_t'][idx],
                'Hoisting_Budget_gpt': data_dict['budget_grade'][idx],
                'Hoisted_Budget_kg': data_dict['budget_gold'][idx]
            }
            records.append(rec)

        result_df = pd.DataFrame(records)
        return result_df
