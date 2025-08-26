"""
PLANT Sheet Data Processor
Extracts daily actuals and budgets for crushed, milled, CIL feed and tails
metrics. The sheet layout varies slightly from month to month—particularly
whether the day values begin in column H or I—but the processor detects
the "Shifts" header to determine where the day numbers start. It then
reads the corresponding rows labelled with metric names to build a flat
table. Missing values (either budgets or actuals) are replaced with 0.

The resulting DataFrame contains one record per day with columns for
actual and budget figures for each metric.
"""

import pandas as pd
from datetime import datetime
import logging
from typing import Dict, List, Optional
from openpyxl import load_workbook

# Reuse a basic logger from the hoisting processor if available
try:
    from hoisting_processor import setup_logger, clean_value
except Exception:
    # Define local fallbacks if imported in isolation
    def setup_logger(name: str, log_file: str, level: int = logging.INFO):
        logger = logging.getLogger(name)
        if logger.handlers:
            return logger
        logger.setLevel(level)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        ch = logging.StreamHandler()
        ch.setFormatter(formatter)
        logger.addHandler(ch)
        return logger

    def clean_value(value):
        if value is None:
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0


class PlantProcessor:
    """Processor for PLANT sheet data extraction"""

    def __init__(self, logger: Optional[logging.Logger] = None):
        self.logger = logger or setup_logger('PlantProcessor', 'logs/plant.log')

    def extract_plant_data(self, file_path: str) -> pd.DataFrame:
        """Extract plant data into a flat table with actuals and budgets.

        Parameters
        ----------
        file_path : str
            Path to the Excel report containing a 'PLANT' worksheet.

        Returns
        -------
        pandas.DataFrame
            DataFrame with one record per day and columns capturing
            actual and budget figures for crushed, milled, CIL feed and tails.
        """
        self.logger.info("Loading PLANT sheet...")
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb['PLANT']
        except Exception as e:
            self.logger.error(f"Failed to load PLANT sheet: {e}")
            return pd.DataFrame()

        # Determine the month/year for building proper dates by locating the
        # cell labelled 'Month:' and reading the adjacent cell. If it's not
        # present or not a datetime, month_start_date remains None.
        month_start_date: Optional[datetime] = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower().startswith('month'):
                    # openpyxl's Cell.offset uses positional arguments
                    candidate = cell.offset(0, 1).value
                    if isinstance(candidate, datetime):
                        month_start_date = candidate
                    break
            if month_start_date:
                break

        if month_start_date is None:
            self.logger.warning("Month start date not found in PLANT sheet; dates will be None")

        # Find the 'Shifts' label which precedes the day numbers. This could
        # reside in different columns depending on the file version. We search
        # for it and note its row and column index.
        shifts_row_idx: Optional[int] = None
        shifts_col_idx: Optional[int] = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == 'shifts':
                    shifts_row_idx = cell.row
                    shifts_col_idx = cell.column
                    break
            if shifts_row_idx is not None:
                break

        if shifts_row_idx is None:
            self.logger.error("'Shifts' header not found in PLANT sheet")
            return pd.DataFrame()

        # Day numbers begin in the cell immediately to the right of the
        # 'Shifts' label
        day_start_col = shifts_col_idx + 1
        day_nums: List[int] = []
        col = day_start_col
        # Read consecutive day numbers until encountering a blank cell or non-numeric value
        while True:
            val = ws.cell(row=shifts_row_idx, column=col).value
            if val is None:
                break
            try:
                day_int = int(val)
                day_nums.append(day_int)
            except Exception:
                break
            col += 1

        if not day_nums:
            self.logger.error("No day numbers found in PLANT sheet")
            return pd.DataFrame()

        # Define the metric labels we expect in the sheet. These strings
        # correspond to the text found in the column immediately preceding the
        # daily data. We map internal names to these sheet labels.
        label_map: Dict[str, str] = {
            'budget_tonnes': 'Budget Tonnes (t)',
            'crushed_tonnes': 'Daily Crushed (t)',
            'milled_tonnes': 'Daily Milled (t)',
            'budget_grade': 'Budget Grade (g/t)',
            'crushed_grade': 'Daily Crushed (g/t)',
            'milled_grade': 'Daily Milled (g/t)',
            'cil_grade': 'Daily CIL (g/t)',
            'budget_tails_grade': 'Budget Tails (g/t)',
            'tails_grade': 'Daily Tails (g/t)',
            'budget_gold': 'Budget Gold (kg)',
            'crushed_gold': 'Daily Crushed (kg)',
            'milled_gold': 'Daily Milled (kg)',
            'cil_gold': 'Daily CIL (kg)',
            'budget_tails_gold': 'Budget Tails (kg)',
            'tails_gold': 'Daily Tails (kg)'
        }

        # Locate the row index for each label. We scan the first 30 rows for
        # performance but extend as necessary. If a label isn't found, its
        # row index remains None and its values will default to zero.
        label_rows: Dict[str, Optional[int]] = {key: None for key in label_map}
        max_search_rows = max(30, ws.max_row)
        for row in ws.iter_rows(min_row=1, max_row=max_search_rows):
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    stripped = val.strip()
                    for key, label in label_map.items():
                        if stripped == label and label_rows[key] is None:
                            label_rows[key] = cell.row
                    # If we've found all labels we can stop early
            if all(idx is not None for idx in label_rows.values()):
                break

        # Prepare storage for data; initialise each metric with zero-filled lists
        data: Dict[str, List[float]] = {}
        for key in label_map:
            data[key] = [0.0] * len(day_nums)

        # Extract values for each label across the day columns
        for key, row_idx in label_rows.items():
            if row_idx is None:
                # Leave as zeros when missing
                continue
            for idx, col_offset in enumerate(range(day_start_col, day_start_col + len(day_nums))):
                value = ws.cell(row=row_idx, column=col_offset).value
                data[key][idx] = clean_value(value)

        # Build date list using the month_start_date if available
        dates: List[Optional[datetime]] = []
        if isinstance(month_start_date, datetime):
            year = month_start_date.year
            month = month_start_date.month
            for d in day_nums:
                try:
                    dates.append(datetime(year, month, int(d)))
                except Exception:
                    dates.append(None)
        else:
            dates = [None] * len(day_nums)

        # Assemble the flat table. For each day we combine actual and budget
        # values according to the business rules: budgets for crushed and milled
        # tonnes share the same 'Budget Tonnes (t)' series; budgets for grade
        # share 'Budget Grade (g/t)' except for tails which use
        # 'Budget Tails (g/t)'; budgets for gold share 'Budget Gold (kg)' except
        # for tails which use 'Budget Tails (kg)'.
        records: List[Dict[str, object]] = []
        for idx, date_val in enumerate(dates):
            rec = {
                'Date': date_val,
                # Crushed metrics
                'Plant_Crushed_Actual_t': data['crushed_tonnes'][idx],
                'Plant_Crushed_Budget_t': data['budget_tonnes'][idx],
                'Plant_Crushed_Actual_gpt': data['crushed_grade'][idx],
                'Plant_Crushed_Budget_gpt': data['budget_grade'][idx],
                'Plant_Crushed_Actual_kg': data['crushed_gold'][idx],
                'Plant_Crushed_Budget_kg': data['budget_gold'][idx],
                # Milled metrics
                'Plant_Milled_Actual_t': data['milled_tonnes'][idx],
                'Plant_Milled_Budget_t': data['budget_tonnes'][idx],
                'Plant_Milled_Actual_gpt': data['milled_grade'][idx],
                'Plant_Milled_Budget_gpt': data['budget_grade'][idx],
                'Plant_Milled_Actual_kg': data['milled_gold'][idx],
                'Plant_Milled_Budget_kg': data['budget_gold'][idx],
                # CIL Feed metrics
                'Plant_CILFeed_Actual_gpt': data['cil_grade'][idx],
                'Plant_CILFeed_Budget_gpt': data['budget_grade'][idx],
                'Plant_CILFeed_Actual_kg': data['cil_gold'][idx],
                'Plant_CILFeed_Budget_kg': data['budget_gold'][idx],
                # Tails metrics
                'Plant_Tails_Actual_gpt': data['tails_grade'][idx],
                'Plant_Tails_Budget_gpt': data['budget_tails_grade'][idx],
                'Plant_Tails_Actual_kg': data['tails_gold'][idx],
                'Plant_Tails_Budget_kg': data['budget_tails_gold'][idx]
            }
            records.append(rec)

        return pd.DataFrame(records)